import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkcalendar import DateEntry
import json, csv, re, sys

# import subprocess
import pandas as pd
from datetime import datetime, timedelta
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from pathlib import Path
import hashlib, uuid
import markdown, webbrowser

__version__ = "1.0.1"


class LicenseChecker:
    def get_pc_id(self):
        """Generate a unique computer ID"""
        mac = uuid.getnode().to_bytes(6, "big").hex()
        return hashlib.sha256(mac.encode()).hexdigest()[:16]

    def check_license(self):
        """License check"""
        try:
            with open("license.key", "r") as f:
                return f.read().strip() == self.get_pc_id()
        except:
            return False


class AutoCompleteEntry(tk.Entry):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self._completion_list = []
        self._hits = []
        self._hit_index = 0

        # Create dropdown list
        self.listbox = tk.Listbox(master, width=self["width"])
        self.listbox.bind("<ButtonRelease-1>", self.on_listbox_select)
        self.listbox.bind("<KeyRelease>", self.on_listbox_keyrelease)
        self.listbox.place_forget()  # Hide the list by default

        self.bind("<KeyRelease>", self.handle_keyrelease)

    def set_completion_list(self, completion_list):
        self._completion_list = sorted(completion_list)

    def handle_keyrelease(self, event):
        if event.keysym in (
            "BackSpace",
            "Left",
            "Right",
            "Up",
            "Down",
            "Return",
            "Tab",
        ):
            self.listbox.place_forget()  # Hide the list
            return

        value = self.get()
        if value == "" or len(value) < 1:
            self._hits = []
            self.listbox.place_forget()  # Hide the list
            return

        # Search for matches anywhere in the string
        self._hits = [
            item for item in self._completion_list if value.lower() in item.lower()
        ]

        # Refresh the Listbox
        self.update_listbox()

    def update_listbox(self):
        self.listbox.delete(0, tk.END)  # Clear the current list
        for item in self._hits:
            self.listbox.insert(tk.END, item)  # Add matching items
        if self._hits:
            # Position the list directly under the input field
            x = self.winfo_x()
            y = self.winfo_y() + self.winfo_height()
            self.listbox.place(x=x, y=y)  # Show the list
            self.listbox.lift()  # Bring the list to the front
        else:
            self.listbox.place_forget()  # Hide the list when there are no matches

    def on_listbox_select(self, event):
        # Fill the entry with the selected list item
        selected = self.listbox.get(self.listbox.curselection())
        self.delete(0, tk.END)
        self.insert(0, selected)
        self.listbox.place_forget()  # Hide the list after selection

    def on_listbox_keyrelease(self, event):
        if event.keysym == "Up":
            self.listbox.selection_clear(0, tk.END)
            if self._hit_index > 0:
                self._hit_index -= 1
            self.listbox.selection_set(self._hit_index)
        elif event.keysym == "Down":
            self.listbox.selection_clear(0, tk.END)
            if self._hit_index < len(self._hits) - 1:
                self._hit_index += 1
            self.listbox.selection_set(self._hit_index)
        elif event.keysym in ("Return", "Tab"):
            self.on_listbox_select(event)


class TimeSelector:
    def __init__(self, parent):
        self.frame = ttk.Frame(parent)
        self.frame.grid()

        # tk.Label(self.frame, text="Select time:").grid(row=0, column=0, padx=5, pady=5, sticky="w")

        tk.Label(self.frame, text="Hours:").grid(
            row=1, column=0, padx=0, pady=5, sticky="e"
        )

        self.hour_combobox = ttk.Combobox(self.frame, values=list(range(24)), width=3)
        self.hour_combobox.grid(row=1, column=1, padx=0, pady=5, sticky="w")
        self.hour_combobox.set("0")

        tk.Label(self.frame, text="Minutes:").grid(
            row=1, column=2, padx=5, pady=5, sticky="w"
        )

        self.minute_combobox = ttk.Combobox(self.frame, values=list(range(60)), width=3)
        self.minute_combobox.grid(row=1, column=3, padx=0, pady=5, sticky="w")
        self.minute_combobox.set("0")

    def get_time(self):
        hour = self.hour_combobox.get()
        minute = self.minute_combobox.get()
        return f"{hour}:{minute}"


class EditableTreeview(ttk.Treeview):
    def __init__(
        self, parent, columns, valid_values=None, update_app=None, *args, **kwargs
    ):
        super().__init__(parent, columns=columns, show="headings", *args, **kwargs)
        self.validator = Validator(valid_values or {})
        self.update_app = update_app

        for col in columns:
            self.heading(
                col, text=col, command=lambda _col=col: self.sort_column(_col, False)
            )
            self.column(col, width=100)

        self.data = []
        self.columns_list = columns
        self.id_counter = 1  # Counter for unique IDs
        self.editing_entry = None
        self.non_editable_columns = ["ID"]  # Columns that cannot be edited
        self.bind("<Double-1>", self.on_double_click)
        # Key bindings
        self.bind("<Return>", self.on_edit_row)
        self.bind("<Delete>", self.delete_row)  # Bind the Delete key

        # Add handlers for keyboard shortcuts
        self.bind("<Control-c>", self.cmd_copy)
        self.bind("<Control-v>", self.cmd_paste)
        self.bind("<Control-x>", self.cmd_cut)
        self.bind("<Control-a>", self.cmd_select_all)

    def keypress(self, e):
        """Handle keyboard shortcuts for paste, copy, and cut"""
        if e.keycode == 86 and e.keysym != "v":
            self.cmd_paste()
        elif e.keycode == 67 and e.keysym != "c":
            self.cmd_copy()
        elif e.keycode == 88 and e.keysym != "x":
            self.cmd_cut()
        elif e.keycode == 65 and e.keysym != "a":
            self.cmd_select_all()

    def cmd_copy(self, event=None):
        """Copy handler"""
        widget = self.focus_get()  # Widget currently focused
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Copy>>")
        elif isinstance(widget, EditableTreeview):
            selected_item = self.selection()
            if selected_item:
                values = self.item(selected_item, "values")
                self.clipboard_clear()
                self.clipboard_append("\t".join(values))

    def cmd_cut(self, event=None):
        """Cut handler"""
        widget = self.focus_get()  # Widget currently focused
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Cut>>")

    def cmd_paste(self, event=None):
        """Paste handler"""
        widget = self.focus_get()  # Widget currently focused
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Paste>>")

    def cmd_select_all(self, event=None):
        """Select-all handler"""
        widget = self.focus_get()  # Widget currently focused
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<SelectAll>>")
        elif isinstance(widget, EditableTreeview):
            self.selection_set(self.get_children())

    def on_double_click(self, event):
        selected_item = self.selection()
        if selected_item:
            column = self.identify_column(event.x)
            col_index = int(column.replace("#", "")) - 1  # Column index
            col_name = self.columns_list[col_index]

            # Check whether the column is editable
            if col_name not in self.non_editable_columns:  # Only allow editable columns
                self.edit_cell(selected_item, col_index)

    def edit_cell(self, item, col):
        x, y, width, height = self.bbox(item, col)
        cell_value = self.item(item, "values")[col]

        entry = tk.Entry(self)
        entry.insert(0, cell_value)
        entry.place(x=x, y=y, width=width, height=height)

        entry.focus()
        entry.bind("<Return>", lambda event: self.save_cell(entry, item, col))
        entry.bind("<FocusOut>", lambda event: self.cancel_edit(entry))
        entry.bind("<Escape>", lambda event: self.cancel_edit(entry))

    def save_cell(self, entry, item, col):
        if entry.winfo_exists():
            new_value = entry.get()
            column = self.columns_list[col]

            if new_value.strip() == "":
                messagebox.showerror("Error", "Value cannot be empty")
                entry.focus_set()
                return

            if self.validator.validate_value(column, new_value):
                self.item(item, values=self.get_values(item, col, new_value))
            entry.destroy()
            if self.update_app:
                self.update_app()

    def cancel_edit(self, entry):
        if entry.winfo_exists():
            entry.destroy()

    def get_values(self, item, col, new_value):
        values = list(self.item(item, "values"))
        values[col] = new_value
        return values

    def get_column_values_by_index(self, column_index):
        # Collect all values from the specified column by its index
        column_values = [
            self.item(row_id, "values")[column_index] for row_id in self.get_children()
        ]

        return column_values

    def insert_data(self, data):
        self.data = data
        self.populate_data(self.data)

    def clear_data(self):
        self.delete(*self.get_children())

    def populate_data(self, data):
        self.clear_data()
        for row in data:
            self.insert("", "end", values=row)

    def filter_rows(self, filter_text):
        if not filter_text:
            filtered_data = self.data
        else:
            filtered_data = [
                row
                for row in self.data
                if any(filter_text.lower() in str(value).lower() for value in row)
            ]
        self.populate_data(filtered_data)

    def sort_column(self, col, reverse):
        data = [(self.item(k, "values"), k) for k in self.get_children("")]
        data.sort(key=lambda x: x[0][self.columns_list.index(col)], reverse=reverse)

        self.data = [x[0] for x in data]
        self.populate_data(self.data)

        self.heading(col, command=lambda: self.sort_column(col, not reverse))

    def generate_unique_id(self):
        existing_ids = [
            int(self.item(row_id, "values")[0]) for row_id in self.get_children()
        ]

        # Generate a unique ID that is greater than the current maximum
        if existing_ids:
            return max(existing_ids) + 1
        else:
            return 1  # Start from 1 when there are no records

    def add_row(self, row_data):
        # Generate a unique ID
        new_id = self.generate_unique_id()
        row_data_with_id = [new_id] + row_data
        self.data.append(tuple(row_data_with_id))
        self.insert("", "end", values=row_data_with_id)

    def delete_row(self, event=None):
        selected_items = self.selection()
        if selected_items:
            confirm = messagebox.askyesno(
                "Delete Confirmation",
                "Are you sure you want to delete the selected rows?",
            )
            if confirm:
                for item_id in selected_items:
                    item_index = self.index(item_id)
                    if 0 <= item_index < len(self.data):
                        del self.data[item_index]
                    self.delete(item_id)

    def on_edit_row(self, event=None):
        selected_item = self.selection()
        if not selected_item:
            return

        item_id = selected_item[0]
        current_values = list(self.item(item_id, "values"))

        dialog = tk.Toplevel(self)
        dialog.resizable(False, False)
        dialog.title("Edit row")
        entries = []
        dialog.grab_set()

        for i, (value, column) in enumerate(
            zip(current_values[1:], self["columns"][1:])
        ):
            label = tk.Label(dialog, text=column)
            label.grid(row=i, column=0, padx=10, pady=5)
            entry = tk.Entry(dialog)
            entry.insert(0, value)
            # Check whether the column is editable
            if column in self.non_editable_columns:
                entry.config(state="readonly")  # Set to read-only
            entry.grid(row=i, column=1, padx=10, pady=5)
            entries.append(entry)

        def on_ok():
            updated_values = [current_values[0]] + [entry.get() for entry in entries]

            if any(val.strip() == "" for val in updated_values[1:]):  # Skip ID
                messagebox.showerror("Error", "None of the fields can be empty")
                return

            self.item(item_id, values=updated_values)
            item_index = self.index(item_id)
            if 0 <= item_index < len(self.data):
                self.data[item_index] = tuple(updated_values)
            else:
                messagebox.showerror("Error", "Index is out of data list bounds.")
            dialog.destroy()

        ok_button = tk.Button(dialog, text="OK", command=on_ok)
        ok_button.grid(row=len(self["columns"]) - 1, column=0, columnspan=2, pady=10)

        dialog.wait_window(dialog)


class Validator:
    def __init__(self, valid_values):
        self.valid_values = valid_values

    def validate_value(self, col, value):
        """Validation based on valid_values"""
        valid_type = self.valid_values.get(col)

        if isinstance(valid_type, list):
            # Ensure the value is in the allowed list
            if value not in valid_type:
                messagebox.showerror(
                    "Error", f"Value must be one of: {','.join(valid_type)}"
                )
                return False

        elif valid_type == "datetime_format":
            # Validate date and time in DD.MM.YYYY HH:MM format
            try:
                datetime.strptime(value, "%d.%m.%Y %H:%M")
            except ValueError:
                messagebox.showerror(
                    "Error", "Invalid date format. Expected: DD.MM.YYYY HH:MM"
                )
                return False

        elif valid_type == "positive_decimal":
            # Validate positive decimal numbers (both dot and comma)
            if value and not re.match(r"^\d*([.,]?\d+)?$", value):
                messagebox.showerror("Error", "Value must be a positive number")
                return False

        elif valid_type == "positive_integer":
            # Validate positive integers (without blocking intermediate states)
            if value and (not value.isdigit() or int(value) <= 0):
                messagebox.showerror("Error", "Value must be a positive integer")
                return False

        return True


class SettingsWindow(tk.Toplevel):
    def __init__(self, parent, parent_app):
        super().__init__(parent)
        self.parent_app = parent_app
        self.title("Settings")
        self.geometry("420x300")

        self.create_widgets()
        self.load_settings()

    def create_widgets(self):
        """Create widgets in the settings window"""
        button_width = 15  # Button width in characters

        # Add row and column weights
        self.grid_rowconfigure(1, weight=1)  # Machines list
        self.grid_rowconfigure(4, weight=0)  # Buttons are fixed

        # Helper function for creating labels and entries
        def create_label_and_entry(label_text, row, col):
            tk.Label(self, text=label_text).grid(
                row=row, column=col, padx=10, pady=5, sticky="w"
            )
            entry = tk.Entry(self)
            entry.grid(row=row, column=col + 1, padx=10, pady=5, sticky="e")
            return entry

        # Add a new machine
        tk.Label(self, text="Add machine:").grid(
            row=0, column=0, padx=10, pady=5, sticky="w"
        )
        self.new_machine_entry = tk.Entry(self)
        self.new_machine_entry.grid(row=0, column=1, padx=10, pady=5, sticky="e")
        self.add_button = tk.Button(self, text="Add", command=self.add_machine)
        self.add_button.config(width=button_width)
        self.add_button.grid(row=0, column=2, padx=5, pady=5, sticky="e")

        # Machines list
        tk.Label(self, text="Machines list:").grid(
            row=1, column=0, padx=10, pady=5, sticky="w"
        )
        self.machine_listbox = tk.Listbox(self, height=5)
        self.machine_listbox.grid(row=1, column=1, padx=10, pady=5, sticky="nsew")
        self.remove_button = tk.Button(self, text="Delete", command=self.remove_machine)
        self.remove_button.config(width=button_width)
        self.remove_button.grid(row=1, column=2, padx=5, pady=5, sticky="e")

        # Style
        tk.Label(self, text="Choose style:").grid(
            row=2, column=0, padx=10, pady=5, sticky="w"
        )
        self.style_combo = ttk.Combobox(
            self,
            width=10,
            values=["default", "clam", "alt", "classic", "vista", "xpnative"],
        )
        self.style_combo.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

        # CSV separator
        self.csv_separator_entry = create_label_and_entry("CSV separator:", 3, 0)
        self.csv_separator_entry.config(width=5)

        # Encoding
        tk.Label(self, text="Choose encoding:").grid(
            row=4, column=0, padx=10, pady=5, sticky="w"
        )
        self.encoding_combo = ttk.Combobox(
            self, width=10, values=["windows-1251", "utf-8"]
        )
        self.encoding_combo.grid(row=4, column=1, padx=10, pady=5, sticky="ew")

        # Buttons at the bottom of the form
        button_frame = tk.Frame(self)
        button_frame.grid(row=5, column=0, columnspan=3, padx=10, pady=10, sticky="ew")

        self.default_button = tk.Button(
            button_frame, text="Default", command=self.load_default_settings
        )
        self.default_button.config(width=button_width)
        self.default_button.pack(side=tk.LEFT, padx=5)

        self.save_button = tk.Button(
            button_frame, text="OK", command=self.save_settings
        )
        self.save_button.config(width=button_width)
        self.save_button.pack(side=tk.LEFT, padx=5)

        self.close_button = tk.Button(button_frame, text="Cancel", command=self.destroy)
        self.close_button.config(width=button_width)
        self.close_button.pack(side=tk.LEFT, padx=5)

    def load_settings(self):
        settings_path = Path("settings.json").resolve()
        # Check whether the file exists and is not empty
        if settings_path.exists() and settings_path.stat().st_size > 0:
            try:
                with settings_path.open("r", encoding="utf-8") as f:
                    settings = json.load(f)
            except json.JSONDecodeError as e:
                print(f"JSON decode error: {e}")

            # Fill the fields
            self.style_combo.set(settings.get("style", "vista"))
            self.csv_separator_entry.delete(0, tk.END)  # Clear previous value
            self.csv_separator_entry.insert(0, settings.get("csv_separator", ";"))
            self.encoding_combo.set(settings.get("encoding", "windows-1251"))

            # Load machines list
            self.parent_app.all_machines = settings.get("machines", [])  # Machine list
            self.update_machine_listbox()  # Refresh machines display

        else:
            print("Settings file does not exist or is empty.")
            self.load_default_settings()  # Load defaults

    def load_default_settings(self):
        """Load default settings"""
        self.style_combo.set("vista")
        self.csv_separator_entry.delete(0, tk.END)
        self.csv_separator_entry.insert(0, ";")
        self.encoding_combo.set("windows-1251")
        self.parent_app.all_machines = [
            "HAAS VF-3",
            "DMU-50-1",
            "DMU-50-2",
            "DMU-70",
            "DMU-75",
            "DMG-M1-1",
            "DMG-M1-2",
            "DMG-M1-3",
            "DMC-835",
        ]
        self.update_machine_listbox()

    def save_settings(self):
        """Save settings to a JSON file"""
        style = self.style_combo.get()
        csv_separator = self.csv_separator_entry.get().strip()
        encoding = self.encoding_combo.get()

        if not csv_separator:
            messagebox.showerror("Error", "CSV separator cannot be empty!")
            return

        settings = {
            "style": style,
            "csv_separator": csv_separator,
            "encoding": encoding,
            "machines": self.parent_app.all_machines,
        }

        settings_path = Path("settings.json").resolve()
        with settings_path.open("w", encoding="utf-8") as f:
            json.dump(settings, f)

        self.parent_app.change_style(style)
        self.parent_app.csv_separator = csv_separator
        self.parent_app.encoding = encoding

        self.destroy()

    def add_machine(self):
        """Add a new machine"""
        new_machine = self.new_machine_entry.get().strip()
        if new_machine:
            if new_machine not in self.parent_app.all_machines:
                self.parent_app.all_machines.append(new_machine)
                self.parent_app.machine_combo.config(
                    values=self.parent_app.all_machines
                )
                self.parent_app.machine_combo2.config(
                    values=self.parent_app.all_machines
                )
                self.update_machine_listbox()
                self.new_machine_entry.delete(0, tk.END)
            else:
                messagebox.showwarning("Error", "This machine is already added!")
        else:
            messagebox.showwarning("Error", "Enter a machine name!")

    def remove_machine(self):
        """Remove the selected machine"""
        selected = self.machine_listbox.curselection()
        if selected:
            machine = self.machine_listbox.get(selected)
            self.parent_app.all_machines.remove(machine)
            self.update_machine_listbox()

    def update_machine_listbox(self):
        """Refresh the Listbox contents"""
        self.machine_listbox.delete(0, tk.END)
        for machine in self.parent_app.all_machines:
            self.machine_listbox.insert(tk.END, machine)


class App:
    def __init__(self, root):
        self.root = root
        self.setup_window()

        # # License check
        # if not LicenseChecker().check_license():
        #     messagebox.showerror(
        #         "Error", "The application is not activated for this computer"
        #     )
        #     root.destroy()
        #     return

        self.load_settings()
        self.setup_variables()
        self.create_widgets()

        self.load_database_silently(self.current_file_path, self.csv_separator)
        self.update_statusbar()

    """Update"""

    def update_app(self):
        self.mark_data_as_modified()
        self.update_statusbar()
        self.highlight_conflicts()
        self.update_gantt_chart()

    def update_gantt_chart(self, event=None):
        now = datetime.now()
        range_selection = self.range_combo.get()
        if range_selection == "Day":
            start_of_day = datetime(
                now.year, now.month, now.day
            )  # Start of the current day
            end_of_day = (
                start_of_day + timedelta(days=1) - timedelta(seconds=1)
            )  # End of the current day
            min_dt, max_dt = start_of_day, end_of_day
        elif range_selection == "Week":
            start_of_week = self.get_monday_at_midnight(
                now
            )  # Start of the week (Monday)
            end_of_week = start_of_week + timedelta(days=7)  # End of the week
            min_dt, max_dt = start_of_week, end_of_week
        else:  # For the month
            start_of_month = datetime(now.year, now.month, 1)
            if now.month == 12:
                end_of_month = datetime(now.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_of_month = datetime(now.year, now.month + 1, 1) - timedelta(days=1)
            min_dt, max_dt = start_of_month, end_of_month

        if not self.task_table.get_children():
            self.ax.clear()
            self.ax.axis("off")
            self.canvas.draw()
            return

        self.ax.clear()
        self.ax.axis("on")

        # Generate enough colors for the tasks
        num_tasks = len(self.task_table.get_children())
        colors = self.generate_colors(num_tasks)

        # for item in self.task_table.get_children():
        for index, item in enumerate(self.task_table.get_children()):
            (
                id,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            start_dt = datetime.strptime(start_datetime, "%d.%m.%Y %H:%M")
            end_dt = datetime.strptime(end_datetime, "%d.%m.%Y %H:%M")
            machine_index = self.all_machines.index(machine)

            # Set bar length based on the selected range
            bar_length = (end_dt - start_dt).total_seconds() / (60 * 60 * 24)  # In days

            # print(f"start_dt: {start_dt}, bar_length: {bar_length}, start_offset: {start_offset}")

            # Draw a bar with a single color per machine (kept for reference)
            # self.ax.barh(machine_index, bar_length, left=start_offset, color=plt.cm.tab10(machine_index % 10),
            #             edgecolor="black", label=detail)

            # Draw a bar with an individual color per task
            color = colors[index]  # Pick a color
            self.ax.barh(
                machine_index, bar_length, left=start_dt, color=color, label=detail
            )

            # min_dt = min(min_dt, start_dt)
            # max_dt = max(max_dt, end_dt)

        # Update X axis bounds
        self.ax.set_xlim([min_dt, max_dt])

        # Set X axis ticks
        if range_selection == "Day":
            self.ax.xaxis.set_major_locator(mdates.HourLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        elif range_selection == "Week":
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
        elif range_selection == "Month":
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))

        self.ax.set_xlabel("")
        self.ax.set_yticks(range(len(self.all_machines)))
        self.ax.set_yticklabels(self.all_machines)
        self.ax.xaxis.set_visible(True)
        self.ax.xaxis.grid(False)
        self.ax.yaxis.grid(False)
        self.ax.spines["top"].set_visible(False)
        self.ax.spines["right"].set_visible(False)
        self.ax.spines["left"].set_visible(False)
        self.ax.spines["bottom"].set_visible(False)
        self.ax.legend(loc="upper left", bbox_to_anchor=(1, 1), fontsize="small")
        plt.setp(
            self.ax.get_xticklabels(), rotation=90, ha="right", rotation_mode="anchor"
        )
        self.canvas.draw()

    def update_statusbar(self):
        """Update the status bar with counters and shortened path"""
        num_tasks = len(self.task_table.get_children())
        num_items = len(self.nomenclature_table.get_children())
        if self.current_file_path == "":
            short_path = "Not set"
        else:
            short_path = self.shorten_path(self.current_file_path)

        self.statusbar.config(
            text=f"Total tasks: {num_tasks}  | Total records in database: {num_items} | Database: {short_path}"
        )

    def update_detail_list(self):
        self.detailList = self.nomenclature_table.get_column_values_by_index(1)
        self.detail_entry.set_completion_list(self.detailList)
        self.mark_data_as_modified()
        self.update_statusbar()

    def mark_data_as_modified(self, *args):
        self.is_data_modified = True

    def highlight_conflicts(self):
        if not self.task_table.get_children():
            return
        self.task_table.tag_configure("conflict", background="red", foreground="white")

        # Remove conflict tags before re-checking
        for item in self.task_table.get_children():
            self.task_table.item(item, tags="")

        machine_tasks = {}

        for item in self.task_table.get_children():
            (
                id,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            start_dt = datetime.strptime(start_datetime, "%d.%m.%Y %H:%M")
            end_dt = datetime.strptime(end_datetime, "%d.%m.%Y %H:%M")

            # Group tasks by machine
            if machine not in machine_tasks:
                machine_tasks[machine] = []
            machine_tasks[machine].append((item, start_dt, end_dt))

        # Check overlaps per machine
        for machine, tasks in machine_tasks.items():
            for i in range(len(tasks)):
                item1, start1, end1 = tasks[i]
                for j in range(i + 1, len(tasks)):
                    item2, start2, end2 = tasks[j]
                    overlap_start = max(start1, start2)
                    overlap_end = min(end1, end2)
                    overlap_duration = (
                        overlap_end - overlap_start
                    ).total_seconds() / 60
                    if overlap_duration >= 0.5:  # Duration in minutes
                        self.task_table.item(item1, tags="conflict")
                        self.task_table.item(item2, tags="conflict")

    """Settings"""

    def load_settings(self):
        """Load settings from a JSON file"""
        settings = {}
        settings_path = Path("settings.json").resolve()

        # Check whether the file exists and is not empty
        if settings_path.exists() and settings_path.stat().st_size > 0:
            try:
                with settings_path.open("r", encoding="utf-8") as f:
                    settings = json.load(f)
            except json.JSONDecodeError as e:
                print(f"JSON decode error: {e}")
        else:
            print("Settings file does not exist or is empty.")

        self.current_style = settings.get("style", "vista")
        self.change_style(self.current_style)
        self.csv_separator = settings.get("csv_separator", ";")
        self.encoding = settings.get("encoding", "windows-1251")
        self.all_machines = settings.get(
            "machines",
            [
                "HAAS VF-3",
                "DMU-50-1",
                "DMU-50-2",
                "DMU-70",
                "DMU-75",
                "DMG-M1-1",
                "DMG-M1-2",
                "DMG-M1-3",
                "DMC-835",
            ],
        )

        geometry = settings.get("window_geometry", None)
        if geometry:
            self.root.geometry(geometry)

        window_state = settings.get("window_state", "normal")
        self.root.state(window_state)

        self.current_file_path = settings.get("last_opened_file", "")
        self.save_settings()

    def save_settings(self):
        """Save settings to a JSON file"""
        geometry = self.root.geometry()
        window_state = self.root.state()

        if self.current_file_path and Path(self.current_file_path).exists():
            path = str(self.current_file_path)
        else:
            path = None

        settings = {
            "csv_separator": self.csv_separator,
            "encoding": self.encoding,
            "machines": self.all_machines,
            "style": self.current_style,
            "window_geometry": geometry,
            "window_state": window_state,
            "last_opened_file": path,
        }

        settings_path = Path("settings.json").resolve()
        with settings_path.open("w", encoding="utf-8") as f:
            json.dump(settings, f)

    def change_style(self, style_name):
        """Change the application style"""
        style = ttk.Style()
        style.theme_use(style_name)
        self.current_style = style_name

    def setup_window(self):
        self.root.title("Planning")
        self.root.geometry("700x600")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_variables(self):
        self.is_data_modified = False
        # Create validator instance
        self.validator = Validator(valid_values=self.get_valid_values())
        self.detailList = []
        self.settings_window = None
        self.about_window = None

    """Build the interface"""

    def create_widgets(self):
        self.create_menu()
        self.create_statusbar()
        self.create_tabs()

        # Set default combobox values
        self.machine_combo.set("DMU-50-1")
        self.machine_combo2.set("DMU-50-1")

    def create_menu(self):
        """Create menu"""
        self.menu = tk.Menu(self.root)
        self.root.config(menu=self.menu)

        # File menu
        file_menu = tk.Menu(self.menu, tearoff=0)
        file_menu.add_command(
            label="Open database", command=self.open_database, accelerator="Ctrl+O"
        )
        file_menu.add_command(
            label="Save database", command=self.save_to_database, accelerator="Ctrl+S"
        )
        file_menu.add_command(label="Close database", command=self.close_database)
        file_menu.add_separator()
        file_menu.add_command(
            label="Import tasks from Excel", command=self.import_tasks
        )
        file_menu.add_command(label="Export tasks to Excel", command=self.export_tasks)
        file_menu.add_command(
            label="Export chart to Excel", command=self.export_diagram
        )
        file_menu.add_separator()
        file_menu.add_command(label="Settings", command=self.open_settings)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.on_close, accelerator="Ctrl+Q")
        self.menu.add_cascade(label="File", menu=file_menu)

        # Edit menu
        edit_menu = tk.Menu(self.menu, tearoff=0)
        edit_menu.add_command(label="Cut", command=self.cmd_cut, accelerator="Ctrl+X")
        edit_menu.add_command(label="Copy", command=self.cmd_copy, accelerator="Ctrl+C")
        edit_menu.add_command(
            label="Paste", command=self.cmd_paste, accelerator="Ctrl+V"
        )
        edit_menu.add_command(
            label="Select all", command=self.cmd_select_all, accelerator="Ctrl+A"
        )
        self.menu.add_cascade(label="Edit", menu=edit_menu)

        # Help menu
        help_menu = tk.Menu(self.menu, tearoff=0)
        help_menu.add_command(label="About", command=self.show_about, accelerator="F1")
        help_menu.add_command(label="User guide", command=self.open_user_guide)
        self.menu.add_cascade(label="Help", menu=help_menu)

        # Bind hotkeys
        self.root.bind_all("<Control-o>", lambda event: self.open_database())
        self.root.bind_all("<Control-s>", lambda event: self.save_to_database())
        self.root.bind_all("<Control-q>", lambda event: self.on_close())
        self.root.bind_all("<F1>", lambda event: self.show_about())
        root.bind("<Control-KeyPress>", self.keypress)

    def keypress(self, e):
        """Handle shortcuts for paste, copy, and cut"""
        if e.keycode == 86 and e.keysym != "v":
            self.cmd_paste()
        elif e.keycode == 67 and e.keysym != "c":
            self.cmd_copy()
        elif e.keycode == 88 and e.keysym != "x":
            self.cmd_cut()
        elif e.keycode == 65 and e.keysym != "a":
            self.cmd_select_all()

    def cmd_copy(self):
        """Copy handler"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Copy>>")

    def cmd_cut(self):
        """Cut handler"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Cut>>")

    def cmd_paste(self):
        """Paste handler"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Paste>>")

    def cmd_select_all(self):
        """Select-all handler"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<SelectAll>>")

    def create_tabs(self):
        """Create tabs"""
        self.tab_control = ttk.Notebook(self.root)
        self.tab_planning = ttk.Frame(self.tab_control)
        self.tab_diagram = ttk.Frame(self.tab_control)
        self.tab_nomenclature = ttk.Frame(self.tab_control)
        self.current_item = None
        self.tab_control.add(self.tab_planning, text="Planning")
        self.tab_control.add(self.tab_diagram, text="Diagram")
        self.tab_control.add(self.tab_nomenclature, text="Nomenclature")
        self.tab_control.pack(expand=1, fill="both")

        self.create_planning_tab()
        self.create_diagram_tab()
        self.create_nomenclature_tab()

    def create_statusbar(self):
        """Create the status bar"""
        self.statusbar = ttk.Label(self.root, text="", anchor=tk.W)
        self.statusbar.pack(side=tk.BOTTOM, fill=tk.X)

    def create_planning_tab(self):
        # Planning tab
        frame = ttk.Frame(self.tab_planning)
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Column width setup
        frame.grid_columnconfigure(0, minsize=20)
        frame.grid_columnconfigure(1, minsize=50)

        # Input fields
        tk.Label(frame, text="Part").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        # self.detail_entry = tk.Entry(frame, width=100)
        self.detail_entry = AutoCompleteEntry(frame, width=100)
        self.detail_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Setup").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.setup_entry = tk.Entry(frame, width=13)
        self.setup_entry.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "setup")
        )
        self.setup_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Machine").grid(
            row=2, column=0, padx=5, pady=5, sticky="w"
        )
        self.machine_combo = ttk.Combobox(frame, width=10, values=self.all_machines)
        self.machine_combo.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Quantity, pcs").grid(
            row=3, column=0, padx=5, pady=5, sticky="w"
        )
        self.quantity_entry = tk.Entry(frame, width=13)
        self.quantity_entry.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "quantity")
        )
        self.quantity_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Time per unit, min").grid(
            row=4, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_entry = tk.Entry(frame, width=13)
        self.time_entry.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "time")
        )
        self.time_entry.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Batch setup time, min").grid(
            row=5, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_setup = tk.Entry(frame, width=13)
        self.time_setup.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "time")
        )
        self.time_setup.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Start date").grid(
            row=6, column=0, padx=5, pady=5, sticky="w"
        )
        self.start_date = DateEntry(frame, width=10, date_pattern="dd.MM.yyyy")
        self.start_date.grid(row=6, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Start time").grid(
            row=7, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_selector = TimeSelector(frame)
        self.time_selector.frame.grid(row=7, column=1, padx=5, pady=5, sticky="w")

        # Task controls
        control_frame = ttk.Frame(frame)
        control_frame.grid(row=8, columnspan=3, pady=10)
        self.add_task_button = tk.Button(
            control_frame, width=15, text="Query DB", command=self.add_time
        )
        self.add_task_button.pack(side="left", padx=5)
        self.add_task_button = tk.Button(
            control_frame, width=15, text="Add", command=self.add_task
        )
        self.add_task_button.pack(side="left", padx=5)
        self.edit_task_button = tk.Button(
            control_frame, width=15, text="Edit", command=self.edit_task
        )
        self.edit_task_button.pack(side="left", padx=5)
        self.delete_task_button = tk.Button(
            control_frame, width=15, text="Delete", command=self.delete_task
        )
        self.delete_task_button.pack(side="left", padx=5)
        self.clear_tasks_button = tk.Button(
            control_frame, width=15, text="Clear", command=self.clear_tasks
        )
        self.clear_tasks_button.pack(side="left", padx=5)

        # Table
        self.task_table = EditableTreeview(
            frame,
            columns=[
                "ID",
                "Part",
                "Setup",
                "Machine",
                "Qty",
                "Time/unit",
                "Start date",
                "End date",
            ],
            valid_values=self.get_valid_values(),
            update_app=self.update_app,
        )
        self.task_table.grid(row=9, columnspan=3, pady=10, sticky="nsew")
        self.task_table.column(
            "ID", anchor="center", width=1, minwidth=1, stretch=False
        )
        self.task_table.column(
            "Part", anchor="center", width=150, minwidth=100, stretch=True
        )
        self.task_table.column(
            "Setup", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.task_table.column(
            "Machine", anchor="center", width=50, minwidth=50, stretch=True
        )
        self.task_table.column(
            "Qty", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.task_table.column(
            "Time/unit", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.task_table.column(
            "Start date", anchor="center", width=80, minwidth=80, stretch=True
        )
        self.task_table.column(
            "End date", anchor="center", width=80, minwidth=80, stretch=True
        )
        self.task_table.non_editable_columns = [
            "ID",
            "Qty",
            "Time/unit",
            "Start date",
            "End date",
        ]

        # Scrollbar
        scrollbar_y = ttk.Scrollbar(
            frame, orient="vertical", command=self.task_table.yview
        )
        scrollbar_y.grid(row=9, column=3, sticky="ns")
        self.task_table.configure(yscrollcommand=scrollbar_y.set)

        # Filter field
        self.filter_entry = tk.Entry(frame, width=30)
        self.filter_entry.grid(row=10, column=1, sticky="e")
        self.filter_entry.bind(
            "<KeyRelease>",
            lambda event: self.task_table.filter_rows(self.filter_entry.get()),
        )

        # Expand table
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(9, weight=1)

    def create_diagram_tab(self):
        canvas_frame = ttk.Frame(self.tab_diagram)
        canvas_frame.pack(fill="both", expand=True)

        # Add a dropdown for range selection
        self.range_combo = ttk.Combobox(canvas_frame, values=["Day", "Week", "Month"])
        self.range_combo.pack(side="bottom", pady=10)
        self.range_combo.current(1)  # Week by default
        self.range_combo.bind(
            "<<ComboboxSelected>>", self.update_gantt_chart
        )  # Handle selection changes

        # Create figure and axes for the chart
        self.figure = Figure(figsize=(8, 6), dpi=100)
        self.ax = self.figure.add_subplot(111)

        # Create canvas for the chart
        self.canvas = FigureCanvasTkAgg(self.figure, master=canvas_frame)

        # Create navigation toolbar
        self.toolbar = NavigationToolbar2Tk(self.canvas, canvas_frame)
        self.toolbar.update()
        self.toolbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.get_tk_widget().pack(
            side=tk.TOP, fill="both", expand=True
        )  # Center the canvas

        # Bind scroll event
        self.canvas.mpl_connect("scroll_event", self.zoom)

        # Initial chart render
        self.update_gantt_chart()

    def create_nomenclature_tab(self):
        # Nomenclature tab
        frame = ttk.Frame(self.tab_nomenclature)
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Column width setup
        frame.grid_columnconfigure(0, minsize=20)
        frame.grid_columnconfigure(1, minsize=50)

        # Input fields
        tk.Label(frame, text="Part type").grid(
            row=0, column=0, padx=5, pady=5, sticky="w"
        )
        self.part_type_entry = tk.Entry(frame, width=30)
        self.part_type_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Drawing number").grid(
            row=1, column=0, padx=5, pady=5, sticky="w"
        )
        self.drawing_number_entry = tk.Entry(frame, width=30)
        self.drawing_number_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Setup").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.setup_entry2 = tk.Entry(frame, width=13)
        self.setup_entry2.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "setup")
        )
        self.setup_entry2.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Machine").grid(
            row=4, column=0, padx=5, pady=5, sticky="w"
        )
        self.machine_combo2 = ttk.Combobox(frame, width=10, values=self.all_machines)
        self.machine_combo2.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Time per unit, min").grid(
            row=5, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_entry2 = tk.Entry(frame, width=13)
        self.time_entry2.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "time")
        )
        self.time_entry2.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        # Nomenclature controls
        control_frame = ttk.Frame(frame)
        control_frame.grid(row=6, columnspan=3, pady=10)
        self.add_nomenclature_button = tk.Button(
            control_frame, width=15, text="Add", command=self.add_nomenclature
        )
        self.add_nomenclature_button.pack(side="left", padx=5)
        self.edit_nomenclature_button = tk.Button(
            control_frame, width=15, text="Edit", command=self.edit_nomenclature
        )
        self.edit_nomenclature_button.pack(side="left", padx=5)
        self.delete_nomenclature_button = tk.Button(
            control_frame, width=15, text="Delete", command=self.delete_nomenclature
        )
        self.delete_nomenclature_button.pack(side="left", padx=5)

        # Table
        self.nomenclature_table = EditableTreeview(
            frame,
            columns=["ID", "Part", "Setup", "Machine", "Time/unit"],
            valid_values=self.get_valid_values(),
            update_app=self.update_app,
        )
        self.nomenclature_table.grid(row=7, columnspan=3, pady=10, sticky="nsew")
        self.nomenclature_table.column(
            "ID", anchor="center", width=5, minwidth=5, stretch=False
        )
        self.nomenclature_table.column(
            "Part", anchor="center", width=200, minwidth=200, stretch=True
        )
        self.nomenclature_table.column(
            "Setup", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.nomenclature_table.column(
            "Machine", anchor="center", width=50, minwidth=50, stretch=True
        )
        self.nomenclature_table.column(
            "Time/unit", anchor="center", width=10, minwidth=10, stretch=True
        )

        # Scrollbar
        scrollbar_y = ttk.Scrollbar(
            frame, orient="vertical", command=self.nomenclature_table.yview
        )
        scrollbar_y.grid(row=7, column=3, sticky="ns")
        self.nomenclature_table.configure(yscrollcommand=scrollbar_y.set)

        # Filter field
        self.filter_entry2 = tk.Entry(frame, width=30)
        self.filter_entry2.grid(row=8, column=2, sticky="e")
        self.filter_entry2.bind(
            "<KeyRelease>",
            lambda event: self.nomenclature_table.filter_rows(self.filter_entry2.get()),
        )

        # Expand table
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(7, weight=1)

    """Validation"""

    def validate_field(self, field_type, value):
        """Validate fields on focus loss"""
        if field_type in ["setup", "quantity"]:
            col = "Setup" if field_type == "setup" else "Qty"
            return self.validator.validate_value(col, value)
        elif field_type == "time":
            return self.validator.validate_value("Time/unit", value)

        return True

    def on_focus_out(self, event, field_type):
        """Handle focus loss for validation"""
        value = event.widget.get()
        self.validate_field(field_type, value)

    def get_valid_values(self):
        return {
            "Setup": "positive_integer",  # Positive integers
            "Machine": self.all_machines,  # List of machines
            "Qty": "positive_integer",
            "Time/unit": "positive_decimal",  # Positive with a decimal point
            "Start date": "datetime_format",  # Date and time
            "End date": "datetime_format",
        }

    """Menu"""

    def validate_and_import_data(self, data):
        valid_data = []
        errors = []

        for i, row in enumerate(data):
            try:
                # Validate first column (must be an integer)
                int(row[0])

                # Validate 5th column (replace comma and ensure float)
                if len(row) > 4:
                    row[4] = row[4].replace(",", ".")
                    float(row[4])

                # If validation passes, add row to valid data
                valid_data.append(row)
            except Exception as e:
                # On error, record the row index and message
                errors.append((i + 1, str(e)))

        return valid_data, errors

    def load_database_silently(self, file_path=None, delimiter=","):
        if not file_path:
            return

        file_path = Path(file_path) if isinstance(file_path, str) else file_path
        data = []
        try:
            with file_path.open(newline="", encoding=self.encoding) as csvfile:
                reader = csv.reader(csvfile, delimiter=delimiter)
                next(reader)  # Skip header
                for row in reader:
                    if row:
                        data.append(row)
        except FileNotFoundError:
            print(f"File {file_path} not found.")
            return
        except Exception as e:
            print(f"Error loading file: {e}")
            return

        # Validate data before loading into the table
        valid_data, errors = self.validate_and_import_data(data)

        # Output errors to console
        if errors:
            for error in errors:
                print(f"Error in line {error[0]}: {error[1]}")

        if not valid_data:
            print("All rows contain errors.")
            return

        # Load valid data into the TreeView
        self.nomenclature_table.clear_data()
        self.nomenclature_table.insert_data(valid_data)
        self.update_detail_list()
        self.is_data_modified = False
        print(f"Successfully loaded {len(valid_data)} records from {file_path}.")

    def open_database(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not file_path:
            return  # Exit if the user cancels the dialog
        file_path = Path(file_path)

        if self.current_file_path == file_path:
            messagebox.showwarning("Warning", "The file is already open.")
            return

        if not self.csv_separator:
            delimiter = simpledialog.askstring(
                "Choose delimiter", "Enter a delimiter (for example, ',' or ';'):"
            )
        else:
            delimiter = self.csv_separator

        if not delimiter:
            return  # Exit if no delimiter was provided

        data = []
        try:
            with file_path.open(newline="", encoding=self.encoding) as csvfile:
                reader = csv.reader(csvfile, delimiter=delimiter)
                header = next(reader)  # Skip the first row (header)
                for row in reader:
                    if row:  # Ignore empty rows
                        data.append(row)
        except FileNotFoundError:
            messagebox.showerror("Error", "Specified file not found.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {e}")
            return

        if not data:
            messagebox.showwarning(
                "Warning", "The file is empty or incorrectly formatted."
            )
            return

        # Data validation
        valid_data, errors = self.validate_and_import_data(data)

        # Handle validation errors
        if errors:
            error_message = "\n".join(
                [f"Error in line {i}: {error}" for i, error in errors]
            )
            messagebox.showerror(
                "Import errors", f"The following errors were found:\n{error_message}"
            )
            return  # Stop execution when errors are present

        # If validation succeeds, clear the table and add valid data
        self.nomenclature_table.clear_data()
        self.tab_control.select(self.tab_nomenclature)
        self.nomenclature_table.insert_data(valid_data)
        self.current_file_path = file_path
        self.update_detail_list()
        messagebox.showinfo(
            "Success", f"Successfully imported {len(valid_data)} records."
        )

    def export_to_csv(self, file_path):
        file_path = Path(file_path)

        if not self.csv_separator:
            delimiter = simpledialog.askstring(
                "Choose delimiter", "Enter a delimiter (for example, ',' or ';'):"
            )
        else:
            delimiter = self.csv_separator

        if not delimiter:
            return  # Exit if no delimiter was provided
        try:
            with file_path.open("w", newline="", encoding=self.encoding) as csvfile:
                writer = csv.writer(csvfile, delimiter=delimiter)

                # Write headers (if needed)
                writer.writerow(self.nomenclature_table["columns"])

                # Write data from the Treeview
                for row in self.nomenclature_table.get_children():
                    writer.writerow(self.nomenclature_table.item(row)["values"])

            self.current_file_path = file_path
            return True  # Successful export
        except Exception as e:
            messagebox.showerror("Error", f"Could not export data: {e}")
            return False  # Failed export

    def save_to_database(self):
        if self.current_file_path and self.current_file_path.exists():
            success = self.export_to_csv(self.current_file_path)
        else:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv", filetypes=[("CSV files", "*.csv")]
            )
            if file_path:
                success = self.export_to_csv(file_path)
            else:
                return  # Exit if the user cancels the dialog

        if success:
            messagebox.showinfo("Success", "File saved successfully.")
            self.update_detail_list()

    def close_database(self):
        for item in self.nomenclature_table.get_children():
            self.nomenclature_table.delete(item)
        self.current_file_path = ""
        self.update_detail_list()

    def import_tasks(self):
        file_path = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return

        file_path = Path(file_path)
        try:
            # Load data from Excel without transforming dates
            df = pd.read_excel(file_path, dtype=str)

            if df.empty:
                messagebox.showwarning("Error", "The file is empty.")
                return

            self.task_table.clear_data()

            for _, row in df.iterrows():
                # Read values from rows as-is
                start_date = row.get("Start date", "")
                end_date = row.get("End date", "")

                task_data = [
                    row.get("Part", ""),
                    row.get("Setup", ""),
                    row.get("Machine", ""),
                    row.get("Qty", ""),
                    row.get("Time/unit", ""),
                    start_date,
                    end_date,
                ]
                self.task_table.add_row(task_data)

            self.update_app()
            self.is_data_modified = False
            messagebox.showinfo("Success", "Tasks were loaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Error while loading tasks: {str(e)}")

    def save_to_excel(
        self, file_path, data, sheet_name, include_image=None, column_names=None
    ):
        file_path = Path(file_path)
        try:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Use default indexes when column names are not provided
                df = (
                    pd.DataFrame(data, columns=column_names)
                    if column_names
                    else pd.DataFrame(data)
                )
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Add spacing between header and data
                # worksheet.insert_rows(1)
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=1, column=col_num).alignment = Alignment(
                        horizontal="center"
                    )

                # Set column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]  # Convert to list for processing
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2  # Add padding
                    worksheet.column_dimensions[column[0].column_letter].width = (
                        adjusted_width
                    )

                # Center all cell content
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=worksheet.max_row,
                    min_col=1,
                    max_col=worksheet.max_column,
                ):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

                # Add an image when provided
                if include_image:
                    img = Image(include_image)
                    worksheet.add_image(
                        img, "H1"
                    )  # Replace "H1" with a desired cell for the image

            messagebox.showinfo("Success", "Data saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving to Excel: {str(e)}")

    def export_tasks(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return

        file_path = Path(file_path)
        data = []
        for item in self.task_table.get_children():
            (
                _,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            data.append(
                (
                    detail,
                    setup,
                    machine,
                    quantity,
                    time_per_unit,
                    start_datetime,
                    end_datetime,
                )
            )

        # Provide column names
        column_names = [
            "Part",
            "Setup",
            "Machine",
            "Qty",
            "Time/unit",
            "Start date",
            "End date",
        ]
        self.save_to_excel(
            file_path, data, sheet_name="Tasks", column_names=column_names
        )

    def export_diagram(self):
        excel_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save as",
        )
        if not excel_file:
            messagebox.showwarning("Warning", "No file selected. Export canceled.")
            return

        excel_file = Path(excel_file)
        image_path = "gantt_chart.png"
        self.figure.savefig(image_path, bbox_inches="tight")
        plt.close()

        data = []
        for item in self.task_table.get_children():
            (
                _,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            data.append(
                [
                    detail,
                    setup,
                    machine,
                    quantity,
                    time_per_unit,
                    start_datetime,
                    end_datetime,
                ]
            )

        # Provide column names
        column_names = [
            "Part",
            "Setup",
            "Machine",
            "Qty",
            "Time/unit",
            "Start date",
            "End date",
        ]
        self.save_to_excel(
            excel_file,
            data,
            sheet_name="Diagram",
            include_image=image_path,
            column_names=column_names,
        )

    def open_settings(self):
        if self.settings_window is None or not self.settings_window.winfo_exists():
            self.settings_window = SettingsWindow(self.root, self)
        else:
            self.settings_window.lift()

    def on_close(self):
        self.save_settings()
        if self.is_data_modified:
            if not messagebox.askokcancel(
                "Exit",
                "There are unsaved changes in the database. Do you really want to exit?",
            ):
                return
        self.root.quit()
        self.root.destroy()

    def show_about(self):
        if self.about_window is None or not self.about_window.winfo_exists():
            self.about_window = tk.Toplevel()
            self.about_window.title("About")
            self.about_window.geometry("300x200")
            self.about_window.resizable(False, False)

            # Prevent minimizing the window
            self.about_window.wm_attributes("-toolwindow", True)

            app_info = [
                ("Name: ", "bold"),
                ("Planning\n", "normal"),
                ("Version: ", "bold"),
                (f"{__version__}\n", "normal"),
                ("Author: ", "bold"),
                ("MaestroFusion360\n", "normal"),
                ("Description: ", "bold"),
                (
                    "This application helps manage production tasks, plan their execution, "
                    "and visualize them as a Gantt chart.\n",
                    "normal",
                ),
            ]

            text_widget = tk.Text(self.about_window, wrap="word", height=10, width=40)
            text_widget.pack(padx=10, pady=10, expand=True, fill="both")

            for text, style in app_info:
                if style == "bold":
                    text_widget.insert(tk.END, text, ("bold",))
                else:
                    text_widget.insert(tk.END, text)

            text_widget.tag_config("bold", font=("Arial", 10, "bold"))
            text_widget.config(state=tk.DISABLED)

            close_button = ttk.Button(
                self.about_window, text="Close", command=self.about_window.destroy
            )
            close_button.pack(pady=10)

        else:
            self.about_window.lift()

    def open_user_guide(self):
        """Opening the User Guide"""
        if getattr(sys, 'frozen', False):
            base_path = Path(sys._MEIPASS)
        else:
            base_path = Path(__file__).parent

        file_path = base_path / "docs" / "README_EN.md"

        if file_path.exists():
            # subprocess.Popen(["notepad.exe", str(file_path)])  # Open in Notepad
            self.open_markdown_in_browser(file_path) # Open in a browser
        else:
            messagebox.showerror("Error", f"User guide not found:\n{file_path}")


    """Planning tab actions"""

    def add_time(self):

        detail = self.detail_entry.get().strip()
        setup = self.setup_entry.get().strip()
        machine = self.machine_combo.get().strip()
        search_row = (detail, setup, machine)

        # Check for empty fields
        if not all([detail, setup, machine]):
            messagebox.showwarning(
                "Warning", "Please fill in these fields: Part, Setup, Machine"
            )
            return

        for child in self.nomenclature_table.get_children():
            row_values = self.nomenclature_table.item(child, "values")

            if (
                row_values[1] == search_row[0]
                and row_values[2] == search_row[1]
                and row_values[3] == search_row[2]
            ):
                messagebox.showinfo(
                    "Search result", f"Nomenclature found: {row_values}"
                )
                self.time_entry.delete(0, tk.END)
                self.time_entry.insert(0, row_values[4])
                return

        messagebox.showwarning("Search result", "Nomenclature not found.")

    def add_task(self):
        detail = self.detail_entry.get().strip()
        setup = self.setup_entry.get().strip()
        machine = self.machine_combo.get().strip()
        quantity = self.quantity_entry.get().strip()
        time_per_unit = self.time_entry.get().replace(",", ".")
        time_setup = self.time_setup.get().replace(",", ".")
        start_date_str = self.start_date.get()
        start_time_str = self.time_selector.get_time()

        # Check for empty fields
        if not all([detail, setup, machine, quantity, time_per_unit]):
            messagebox.showwarning("Warning", "Please fill in all fields.")
            return

        try:
            duration = (
                int(quantity) * float(time_per_unit) / 60 + float(time_setup) / 60
            )
            start_datetime = datetime.strptime(
                f"{start_date_str} {start_time_str}", "%d.%m.%Y %H:%M"
            )
            end_datetime = start_datetime + timedelta(hours=duration)

            # Add the task to the table
            self.task_table.add_row(
                [
                    detail,
                    setup,
                    machine,
                    quantity,
                    time_per_unit,
                    start_datetime.strftime("%d.%m.%Y %H:%M"),
                    end_datetime.strftime("%d.%m.%Y %H:%M"),
                ]
            )

            self.update_app()
        except ValueError as ve:
            messagebox.showerror("Error", f"Invalid input: {str(ve)}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def edit_task(self):
        self.task_table.on_edit_row()
        self.update_app()

    def delete_task(self):
        self.task_table.delete_row()
        self.update_app()

    def clear_tasks(self):
        for item in self.task_table.get_children():
            self.task_table.delete(item)
        self.update_app()

    """Nomenclature tab actions"""

    def add_nomenclature(self):
        type_detail = self.part_type_entry.get().strip()
        drawing_number = self.drawing_number_entry.get().strip()
        setup = self.setup_entry2.get().strip()
        machine = self.machine_combo2.get().strip()
        time_per_unit = self.time_entry2.get().strip().replace(",", ".")

        if type_detail and drawing_number and setup and machine and time_per_unit:
            detail = f"{type_detail} {drawing_number}"

            # Check for an existing record with the same detail, setup, and machine
            for row_id in self.nomenclature_table.get_children():
                values = self.nomenclature_table.item(row_id, "values")
                existing_detail = values[1].strip()
                existing_setup = values[2].strip()
                existing_machine = values[3].strip()

                if (
                    existing_detail == detail
                    and existing_setup == setup
                    and existing_machine == machine
                ):
                    messagebox.showinfo(
                        "Information",
                        f"A record with these values already exists at position {self.nomenclature_table.index(row_id) + 1}",
                    )

                    # Highlight the existing entry
                    self.nomenclature_table.selection_set(row_id)
                    self.nomenclature_table.see(row_id)
                    return

            # Add a new row when there is no duplicate
            self.nomenclature_table.add_row([detail, setup, machine, time_per_unit])
            self.update_detail_list()
        else:
            messagebox.showwarning("Warning", "Please fill in all fields.")

    def edit_nomenclature(self):
        self.nomenclature_table.on_edit_row()
        self.update_detail_list()

    def delete_nomenclature(self):
        self.nomenclature_table.delete_row()
        self.update_detail_list()

    """Miscellaneous functions and methods"""

    def zoom(self, event):
        xmin, xmax = self.ax.get_xlim()
        xmin_dt = mdates.num2date(xmin)
        xmax_dt = mdates.num2date(xmax)
        zoom_factor = 0.9 if event.step > 0 else 1.1
        range_x = (xmax - xmin) * zoom_factor
        midpoint_x = (xmax + xmin) / 2
        self.ax.set_xlim(midpoint_x - range_x / 2, midpoint_x + range_x / 2)
        new_range_days = (xmax_dt - xmin_dt).days
        if new_range_days <= 1:
            self.ax.xaxis.set_major_locator(mdates.HourLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        elif new_range_days <= 7:
            self.ax.xaxis.set_major_locator(mdates.HourLocator(interval=6))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m %H:%M"))
        elif new_range_days <= 31:
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
        else:
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=5))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m.%Y"))
        self.canvas.draw_idle()

    def generate_colors(self, num_colors):
        """Generate colors in HSV format"""
        return [
            mcolors.hsv_to_rgb((i / num_colors, 0.8, 0.8)) for i in range(num_colors)
        ]

    def get_monday_at_midnight(self, reference_date):
        """Return Monday of the current week with time set to 00:00."""
        days_since_monday = reference_date.weekday()  # Monday - 0, Sunday - 6
        monday = reference_date - timedelta(days=days_since_monday)
        return monday.replace(hour=0, minute=0, second=0, microsecond=0)

    def shorten_path(self, path, max_length=50):

        if path is None:
            return ""

        path = Path(path)

        # If the path is short enough, return it as-is
        if len(str(path)) <= max_length:
            return str(path)

        # Split the path into parts
        parts = list(path.parts)

        # If the path has more than two parts, shorten it
        if len(parts) > 2:
            # Trim the path, replacing middle parts with '...'
            shortened = Path(parts[0], "...", parts[-1])
        else:
            # Otherwise return the original path
            shortened = path

        return str(shortened)

    def open_markdown_in_browser(self, md_path):
        md_file = Path(md_path)
        html_file = md_file.with_suffix(".html")

        # Read the Markdown content
        md_text = md_file.read_text(encoding="utf-8")

        # Convert to HTML
        html = markdown.markdown(md_text, extensions=["fenced_code", "tables"])

        # Add a simple HTML wrapper
        full_html = f"""
        <html>
        <head>
            <meta charset="utf-8">
            <title>{md_file.stem}</title>
            <style>
                body {{ font-family: sans-serif; margin: 40px; max-width: 800px; }}
                code {{ background-color: #eee; padding: 2px 4px; }}
                pre code {{ background-color: #f8f8f8; display: block; padding: 10px; }}
            </style>
        </head>
        <body>{html}</body>
        </html>
        """

        # Save as HTML
        html_file.write_text(full_html, encoding="utf-8")

        # Open in the browser
        webbrowser.open(html_file.as_uri())


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
