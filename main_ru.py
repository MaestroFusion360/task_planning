import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkcalendar import DateEntry
import json, csv, re

# import subprocess
import pandas as pd
from datetime import datetime, timedelta
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from pathlib import Path
import hashlib, uuid
import markdown, webbrowser

__version__ = "0.2.8"


class LicenseChecker:
    """Проверка лицензии"""

    def get_pc_id(self):
        """Генерация уникального ID компьютера"""
        mac = uuid.getnode().to_bytes(6, "big").hex()
        return hashlib.sha256(mac.encode()).hexdigest()[:16]

    def check_license(self):
        """Проверка лицензии"""
        try:
            with open("license.key", "r") as f:
                return f.read().strip() == self.get_pc_id()
        except:
            return False


class AutoCompleteEntry(tk.Entry):
    """Поле ввода с автозаполнением"""

    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self._completion_list = []
        self._hits = []
        self._hit_index = 0

        # Создаем выпадающий список
        self.listbox = tk.Listbox(master, width=self["width"])
        self.listbox.bind("<ButtonRelease-1>", self.on_listbox_select)
        self.listbox.bind("<KeyRelease>", self.on_listbox_keyrelease)
        self.listbox.place_forget()  # Скрываем список по умолчанию

        self.bind("<KeyRelease>", self.handle_keyrelease)

    def set_completion_list(self, completion_list):
        self._completion_list = sorted(completion_list)

    def handle_keyrelease(self, event):
        if event.keysym in (
            "BackSpace",
            "Left",
            "Right",
            "Up",
            "Down",
            "Return",
            "Tab",
        ):
            self.listbox.place_forget()  # Скрываем список
            return

        value = self.get()
        if value == "" or len(value) < 1:
            self._hits = []
            self.listbox.place_forget()  # Скрываем список
            return

        # Поиск совпадений в любом месте строки
        self._hits = [
            item for item in self._completion_list if value.lower() in item.lower()
        ]

        # Обновляем Listbox
        self.update_listbox()

    def update_listbox(self):
        """Обновляем содержимое Listbox"""
        self.listbox.delete(0, tk.END)  # Очищаем текущий список
        for item in self._hits:
            self.listbox.insert(tk.END, item)  # Добавляем совпадения
        if self._hits:
            # Позиционируем список непосредственно под полем ввода
            x = self.winfo_x()
            y = self.winfo_y() + self.winfo_height()
            self.listbox.place(x=x, y=y)  # Показываем список
            self.listbox.lift()  # Поднимаем список на передний план
        else:
            self.listbox.place_forget()  # Скрываем список, если совпадений нет

    def on_listbox_select(self, event):
        """Обработчик выбора элемента из списка"""
        # Заполняем поле ввода выбранным элементом из списка
        selected = self.listbox.get(self.listbox.curselection())
        self.delete(0, tk.END)
        self.insert(0, selected)
        self.listbox.place_forget()  # Скрываем список после выбора

    def on_listbox_keyrelease(self, event):
        if event.keysym == "Up":
            self.listbox.selection_clear(0, tk.END)
            if self._hit_index > 0:
                self._hit_index -= 1
            self.listbox.selection_set(self._hit_index)
        elif event.keysym == "Down":
            self.listbox.selection_clear(0, tk.END)
            if self._hit_index < len(self._hits) - 1:
                self._hit_index += 1
            self.listbox.selection_set(self._hit_index)
        elif event.keysym in ("Return", "Tab"):
            self.on_listbox_select(event)


class TimeSelector:
    """Выбор времени"""

    def __init__(self, parent):
        self.frame = ttk.Frame(parent)
        self.frame.grid()

        # tk.Label(self.frame, text="Выберите время:").grid(row=0, column=0, padx=5, pady=5, sticky="w")

        tk.Label(self.frame, text="Часы:").grid(
            row=1, column=0, padx=0, pady=5, sticky="e"
        )

        self.hour_combobox = ttk.Combobox(self.frame, values=list(range(24)), width=3)
        self.hour_combobox.grid(row=1, column=1, padx=0, pady=5, sticky="w")
        self.hour_combobox.set("0")

        tk.Label(self.frame, text="Минуты:").grid(
            row=1, column=2, padx=5, pady=5, sticky="w"
        )

        self.minute_combobox = ttk.Combobox(self.frame, values=list(range(60)), width=3)
        self.minute_combobox.grid(row=1, column=3, padx=0, pady=5, sticky="w")
        self.minute_combobox.set("0")

    def get_time(self):
        hour = self.hour_combobox.get()
        minute = self.minute_combobox.get()
        return f"{hour}:{minute}"


class EditableTreeview(ttk.Treeview):
    """Editable Treeview c dополнительными функциями"""

    def __init__(
        self, parent, columns, valid_values=None, update_app=None, *args, **kwargs
    ):
        super().__init__(parent, columns=columns, show="headings", *args, **kwargs)
        self.validator = Validator(valid_values or {})
        self.update_app = update_app

        for col in columns:
            self.heading(
                col, text=col, command=lambda _col=col: self.sort_column(_col, False)
            )
            self.column(col, width=100)

        self.data = []
        self.columns_list = columns
        self.id_counter = 1  # Счётчик для уникальных ID
        self.editing_entry = None
        self.non_editable_columns = ["ID"]  # Запрещённые для редактирования колонки
        self.bind("<Double-1>", self.on_double_click)
        # Привязка клавиш
        self.bind("<Return>", self.on_edit_row)
        self.bind("<Delete>", self.delete_row)  # Привязка клавиши Delete

        # Добавляем обработчики для сочетаний клавиш
        self.bind("<Control-c>", self.cmd_copy)
        self.bind("<Control-v>", self.cmd_paste)
        self.bind("<Control-x>", self.cmd_cut)
        self.bind("<Control-a>", self.cmd_select_all)

    def keypress(self, e):
        """Обработчик комбинаций клавиш для вставки, копирования и вырезания"""
        if e.keycode == 86 and e.keysym != "v":
            self.cmd_paste()
        elif e.keycode == 67 and e.keysym != "c":
            self.cmd_copy()
        elif e.keycode == 88 and e.keysym != "x":
            self.cmd_cut()
        elif e.keycode == 65 and e.keysym != "a":
            self.cmd_select_all()

    def cmd_copy(self, event=None):
        """Обработчик команды копирования"""
        widget = self.focus_get()  # Получаем виджет, на который установлен фокус
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Copy>>")
        elif isinstance(widget, EditableTreeview):
            selected_item = self.selection()
            if selected_item:
                values = self.item(selected_item, "values")
                self.clipboard_clear()
                self.clipboard_append("\t".join(values))

    def cmd_cut(self, event=None):
        """Обработчик команды вырезания"""
        widget = self.focus_get()  # Получаем виджет, на который установлен фокус
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Cut>>")

    def cmd_paste(self, event=None):
        """Обработчик команды вставки"""
        widget = self.focus_get()  # Получаем виджет, на который установлен фокус
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Paste>>")

    def cmd_select_all(self, event=None):
        """Обработчик команды выделения всего текста"""
        widget = self.focus_get()  # Получаем виджет, на который установлен фокус
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<SelectAll>>")
        elif isinstance(widget, EditableTreeview):
            self.selection_set(self.get_children())

    def on_double_click(self, event):
        """Обработчик двойного клика для редактирования ячейки"""
        selected_item = self.selection()
        if selected_item:
            column = self.identify_column(event.x)
            col_index = int(column.replace("#", "")) - 1  # Индекс колонки
            col_name = self.columns_list[col_index]

            # Проверка на редактируемость колонки
            if (
                col_name not in self.non_editable_columns
            ):  # Если колонка не в списке запрещённых
                self.edit_cell(selected_item, col_index)

    def edit_cell(self, item, col):
        """Редактирование ячейки"""
        x, y, width, height = self.bbox(item, col)
        cell_value = self.item(item, "values")[col]

        entry = tk.Entry(self)
        entry.insert(0, cell_value)
        entry.place(x=x, y=y, width=width, height=height)

        entry.focus()
        entry.bind("<Return>", lambda event: self.save_cell(entry, item, col))
        entry.bind("<FocusOut>", lambda event: self.cancel_edit(entry))
        entry.bind("<Escape>", lambda event: self.cancel_edit(entry))

    def save_cell(self, entry, item, col):
        """Сохранение редактируемой ячейки"""
        if entry.winfo_exists():
            new_value = entry.get()
            column = self.columns_list[col]

            if new_value.strip() == "":
                messagebox.showerror("Ошибка", "Значение не может быть пустым")
                entry.focus_set()
                return

            if self.validator.validate_value(column, new_value):
                self.item(item, values=self.get_values(item, col, new_value))
            entry.destroy()
            if self.update_app:
                self.update_app()

    def cancel_edit(self, entry):
        """Отмена редактирования ячейки"""
        if entry.winfo_exists():
            entry.destroy()

    def get_values(self, item, col, new_value):
        """Получаем текущие значения строки и заменяем значение в нужной колонке"""
        values = list(self.item(item, "values"))
        values[col] = new_value
        return values

    def get_column_values_by_index(self, column_index):
        """Собираем все значения из указанной колонки по её индексу"""
        column_values = [
            self.item(row_id, "values")[column_index] for row_id in self.get_children()
        ]

        return column_values

    def insert_data(self, data):
        """Вставка данных в таблицу"""
        self.data = data
        self.populate_data(self.data)

    def clear_data(self):
        """Очистка данных в таблице"""
        self.delete(*self.get_children())

    def populate_data(self, data):
        """Заполнение таблицы данными"""
        self.clear_data()
        for row in data:
            self.insert("", "end", values=row)

    def filter_rows(self, filter_text):
        """Фильтрация строк по тексту"""
        if not filter_text:
            filtered_data = self.data
        else:
            filtered_data = [
                row
                for row in self.data
                if any(filter_text.lower() in str(value).lower() for value in row)
            ]
        self.populate_data(filtered_data)

    def sort_column(self, col, reverse):
        """Сортировка по колонке"""
        data = [(self.item(k, "values"), k) for k in self.get_children("")]
        data.sort(key=lambda x: x[0][self.columns_list.index(col)], reverse=reverse)

        self.data = [x[0] for x in data]
        self.populate_data(self.data)

        self.heading(col, command=lambda: self.sort_column(col, not reverse))

    def generate_unique_id(self):
        existing_ids = [
            int(self.item(row_id, "values")[0]) for row_id in self.get_children()
        ]

        # Генерируем уникальный ID, который больше максимального существующего ID
        if existing_ids:
            return max(existing_ids) + 1
        else:
            return 1  # Если записей еще нет, начинаем с 1

    def add_row(self, row_data):
        """Генерация уникального ID и добавление строки"""
        new_id = self.generate_unique_id()
        row_data_with_id = [new_id] + row_data
        self.data.append(tuple(row_data_with_id))
        self.insert("", "end", values=row_data_with_id)

    def delete_row(self, event=None):
        """Удаление выбранной строки"""
        selected_items = self.selection()
        if selected_items:
            confirm = messagebox.askyesno(
                "Подтверждение удаления",
                "Вы уверены, что хотите удалить выбранные строки?",
            )
            if confirm:
                for item_id in selected_items:
                    item_index = self.index(item_id)
                    if 0 <= item_index < len(self.data):
                        del self.data[item_index]
                    self.delete(item_id)

    def on_edit_row(self, event=None):
        """Редактирование строки"""
        selected_item = self.selection()
        if not selected_item:
            return

        item_id = selected_item[0]
        current_values = list(self.item(item_id, "values"))

        dialog = tk.Toplevel(self)
        dialog.resizable(False, False)
        dialog.title("Редактировать строку")
        entries = []
        dialog.grab_set()

        for i, (value, column) in enumerate(
            zip(current_values[1:], self["columns"][1:])
        ):
            label = tk.Label(dialog, text=column)
            label.grid(row=i, column=0, padx=10, pady=5)
            entry = tk.Entry(dialog)
            entry.insert(0, value)
            # Проверка на редактируемость колонки
            if column in self.non_editable_columns:
                entry.config(state="readonly")  # Устанавливаем только для чтения
            entry.grid(row=i, column=1, padx=10, pady=5)
            entries.append(entry)

        def on_ok():
            updated_values = [current_values[0]] + [entry.get() for entry in entries]

            if any(val.strip() == "" for val in updated_values[1:]):  # Пропускаем ID
                messagebox.showerror("Ошибка", "Ни одно из полей не должно быть пустым")
                return

            self.item(item_id, values=updated_values)
            item_index = self.index(item_id)
            if 0 <= item_index < len(self.data):
                self.data[item_index] = tuple(updated_values)
            else:
                messagebox.showerror("Ошибка", "Индекс вне границ списка данных.")
            dialog.destroy()

        ok_button = tk.Button(dialog, text="OK", command=on_ok)
        ok_button.grid(row=len(self["columns"]) - 1, column=0, columnspan=2, pady=10)

        dialog.wait_window(dialog)


class Validator:
    """Валидация данных"""

    def __init__(self, valid_values):
        self.valid_values = valid_values

    def validate_value(self, col, value):
        """Валидация на основе valid_values"""
        valid_type = self.valid_values.get(col)

        if isinstance(valid_type, list):
            # Проверяем, что значение входит в список допустимых
            if value not in valid_type:
                messagebox.showerror(
                    "Ошибка", f"Значение должно быть одним из: {','.join(valid_type)}"
                )
                return False

        elif valid_type == "datetime_format":
            # Проверяем дату и время в формате DD.MM.YYYY HH:MM
            try:
                datetime.strptime(value, "%d.%m.%Y %H:%M")
            except ValueError:
                messagebox.showerror(
                    "Ошибка", "Неверный формат даты. Ожидается формат: DD.MM.YYYY HH:MM"
                )
                return False

        elif valid_type == "positive_decimal":
            # Проверяем положительные дробные числа (учитываем как точку, так и запятую)
            if value and not re.match(r"^\d*([.,]?\d+)?$", value):
                messagebox.showerror(
                    "Ошибка", "Значение должно быть положительным числом"
                )
                return False

        elif valid_type == "positive_integer":
            # Проверяем положительные целые числа (не блокируем промежуточные состояния)
            if value and (not value.isdigit() or int(value) <= 0):
                messagebox.showerror(
                    "Ошибка", "Значение должно быть положительным целым числом"
                )
                return False

        return True


class SettingsWindow(tk.Toplevel):
    """Окно настроек"""

    def __init__(self, parent, parent_app):
        super().__init__(parent)
        self.parent_app = parent_app
        self.title("Настройки")
        self.geometry("420x300")

        self.create_widgets()
        self.load_settings()

    def create_widgets(self):
        """Создание виджетов в окне настроек"""
        button_width = 15  # ширина кнопок в символах

        # Добавление веса строкам и столбцам
        self.grid_rowconfigure(1, weight=1)  # Список станков
        self.grid_rowconfigure(4, weight=0)  # Кнопки не растягиваются

        # Вспомогательная функция для создания меток и полей ввода
        def create_label_and_entry(label_text, row, col):
            tk.Label(self, text=label_text).grid(
                row=row, column=col, padx=10, pady=5, sticky="w"
            )
            entry = tk.Entry(self)
            entry.grid(row=row, column=col + 1, padx=10, pady=5, sticky="e")
            return entry

        # Добавление нового станка
        tk.Label(self, text="Добавить станок:").grid(
            row=0, column=0, padx=10, pady=5, sticky="w"
        )
        self.new_machine_entry = tk.Entry(self)
        self.new_machine_entry.grid(row=0, column=1, padx=10, pady=5, sticky="e")
        self.add_button = tk.Button(self, text="Добавить", command=self.add_machine)
        self.add_button.config(width=button_width)
        self.add_button.grid(row=0, column=2, padx=5, pady=5, sticky="e")

        # Список станков
        tk.Label(self, text="Список станков:").grid(
            row=1, column=0, padx=10, pady=5, sticky="w"
        )
        self.machine_listbox = tk.Listbox(self, height=5)
        self.machine_listbox.grid(row=1, column=1, padx=10, pady=5, sticky="nsew")
        self.remove_button = tk.Button(
            self, text="Удалить", command=self.remove_machine
        )
        self.remove_button.config(width=button_width)
        self.remove_button.grid(row=1, column=2, padx=5, pady=5, sticky="e")

        # Стиль
        tk.Label(self, text="Выбрать стиль:").grid(
            row=2, column=0, padx=10, pady=5, sticky="w"
        )
        self.style_combo = ttk.Combobox(
            self,
            width=10,
            values=["default", "clam", "alt", "classic", "vista", "xpnative"],
        )
        self.style_combo.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

        # Разделитель CSV
        self.csv_separator_entry = create_label_and_entry("Разделитель CSV:", 3, 0)
        self.csv_separator_entry.config(width=5)

        # Кодировка
        tk.Label(self, text="Выбрать кодировку:").grid(
            row=4, column=0, padx=10, pady=5, sticky="w"
        )
        self.encoding_combo = ttk.Combobox(
            self, width=10, values=["windows-1251", "utf-8"]
        )
        self.encoding_combo.grid(row=4, column=1, padx=10, pady=5, sticky="ew")

        # Кнопки внизу формы
        button_frame = tk.Frame(self)
        button_frame.grid(row=5, column=0, columnspan=3, padx=10, pady=10, sticky="ew")

        self.default_button = tk.Button(
            button_frame, text="По умолчанию", command=self.load_default_settings
        )
        self.default_button.config(width=button_width)
        self.default_button.pack(side=tk.LEFT, padx=5)

        self.save_button = tk.Button(
            button_frame, text="OK", command=self.save_settings
        )
        self.save_button.config(width=button_width)
        self.save_button.pack(side=tk.LEFT, padx=5)

        self.close_button = tk.Button(button_frame, text="Отмена", command=self.destroy)
        self.close_button.config(width=button_width)
        self.close_button.pack(side=tk.LEFT, padx=5)

    def load_settings(self):
        """Загрузка настроек из JSON файла"""
        settings_path = Path("settings.json").resolve()
        # Проверяем, существует ли файл и не пуст ли он
        if settings_path.exists() and settings_path.stat().st_size > 0:
            try:
                with settings_path.open("r", encoding="utf-8") as f:
                    settings = json.load(f)
            except json.JSONDecodeError as e:
                print(f"Ошибка декодирования JSON: {e}")

            # Заполняем поля
            self.style_combo.set(settings.get("style", "vista"))
            self.csv_separator_entry.delete(0, tk.END)  # Удаляем предыдущее значение
            self.csv_separator_entry.insert(0, settings.get("csv_separator", ";"))
            self.encoding_combo.set(settings.get("encoding", "windows-1251"))

            # Загружаем список станков
            self.parent_app.all_machines = settings.get(
                "machines", []
            )  # Получаем список станков
            self.update_machine_listbox()  # Обновляем отображение станков

        else:
            print("Файл настроек не существует или пуст.")
            self.load_default_settings()  # Загружаем настройки по умолчанию

    def load_default_settings(self):
        """Загрузка настроек по умолчанию"""
        self.style_combo.set("vista")
        self.csv_separator_entry.delete(0, tk.END)
        self.csv_separator_entry.insert(0, ";")
        self.encoding_combo.set("windows-1251")
        self.parent_app.all_machines = [
            "HAAS VF-3",
            "DMU-50-1",
            "DMU-50-2",
            "DMU-70",
            "DMU-75",
            "DMG-M1-1",
            "DMG-M1-2",
            "DMG-M1-3",
            "DMC-835",
        ]
        self.update_machine_listbox()

    def save_settings(self):
        """Сохранение настроек в JSON файл"""
        style = self.style_combo.get()
        csv_separator = self.csv_separator_entry.get().strip()
        encoding = self.encoding_combo.get()

        if not csv_separator:
            messagebox.showerror("Ошибка", "Разделитель CSV не может быть пустым!")
            return

        settings = {
            "style": style,
            "csv_separator": csv_separator,
            "encoding": encoding,
            "machines": self.parent_app.all_machines,
        }

        settings_path = Path("settings.json").resolve()
        with settings_path.open("w", encoding="utf-8") as f:
            json.dump(settings, f)

        self.parent_app.change_style(style)
        self.parent_app.csv_separator = csv_separator
        self.parent_app.encoding = encoding

        self.destroy()

    def add_machine(self):
        """Добавление нового станка"""
        new_machine = self.new_machine_entry.get().strip()
        if new_machine:
            if new_machine not in self.parent_app.all_machines:
                self.parent_app.all_machines.append(new_machine)
                self.parent_app.machine_combo.config(
                    values=self.parent_app.all_machines
                )
                self.parent_app.machine_combo2.config(
                    values=self.parent_app.all_machines
                )
                self.update_machine_listbox()
                self.new_machine_entry.delete(0, tk.END)
            else:
                messagebox.showwarning("Ошибка", "Этот станок уже добавлен!")
        else:
            messagebox.showwarning("Ошибка", "Введите название станка!")

    def remove_machine(self):
        """Удаление выбранного станка"""
        selected = self.machine_listbox.curselection()
        if selected:
            machine = self.machine_listbox.get(selected)
            self.parent_app.all_machines.remove(machine)
            self.update_machine_listbox()

    def update_machine_listbox(self):
        """Обновление содержимого списка Listbox"""
        self.machine_listbox.delete(0, tk.END)
        for machine in self.parent_app.all_machines:
            self.machine_listbox.insert(tk.END, machine)


class App:
    """Главный класс приложения"""

    def __init__(self, root):
        self.root = root
        self.setup_window()

        # # Проверка лицензии
        # if not LicenseChecker().check_license():
        #     messagebox.showerror(
        #         "Ошибка", "Программа не активирована для этого компьютера"
        #     )
        #     root.destroy()
        #     return

        self.load_settings()
        self.setup_variables()
        self.create_widgets()

        self.load_database_silently(self.current_file_path, self.csv_separator)
        self.update_statusbar()

    def update_app(self):
        """Обновление приложения"""
        self.mark_data_as_modified()
        self.update_statusbar()
        self.highlight_conflicts()
        self.update_gantt_chart()

    def update_gantt_chart(self, event=None):
        """Обновление диаграммы Ганта"""
        now = datetime.now()
        range_selection = self.range_combo.get()
        if range_selection == "День":
            start_of_day = datetime(now.year, now.month, now.day)  # Начало текущего дня
            end_of_day = (
                start_of_day + timedelta(days=1) - timedelta(seconds=1)
            )  # Конец текущего дня
            min_dt, max_dt = start_of_day, end_of_day
        elif range_selection == "Неделя":
            start_of_week = self.get_monday_at_midnight(
                now
            )  # Начало недели (понедельник)
            end_of_week = start_of_week + timedelta(days=7)  # Конец недели
            min_dt, max_dt = start_of_week, end_of_week
        else:  # Для месяца
            start_of_month = datetime(now.year, now.month, 1)
            if now.month == 12:
                end_of_month = datetime(now.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_of_month = datetime(now.year, now.month + 1, 1) - timedelta(days=1)
            min_dt, max_dt = start_of_month, end_of_month

        if not self.task_table.get_children():
            self.ax.clear()
            self.ax.axis("off")
            self.canvas.draw()
            return

        self.ax.clear()
        self.ax.axis("on")

        # Генерация достаточного количества цветов для задач
        num_tasks = len(self.task_table.get_children())
        colors = self.generate_colors(num_tasks)

        # for item in self.task_table.get_children():
        for index, item in enumerate(self.task_table.get_children()):
            (
                id,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            start_dt = datetime.strptime(start_datetime, "%d.%m.%Y %H:%M")
            end_dt = datetime.strptime(end_datetime, "%d.%m.%Y %H:%M")
            machine_index = self.all_machines.index(machine)

            # Установка длины бара в зависимости от выбранного диапазона
            bar_length = (end_dt - start_dt).total_seconds() / (60 * 60 * 24)  # В днях

            # print(f"start_dt: {start_dt}, bar_length: {bar_length}, start_offset: {start_offset}")

            # Строим бар с обводкой и одного цвета бара для каждого станка
            # self.ax.barh(machine_index, bar_length, left=start_offset, color=plt.cm.tab10(machine_index % 10),
            #             edgecolor="black", label=detail)

            # Строим бар с индивидуальным цветом для задачи
            color = colors[index]  # Выбираем цвет
            self.ax.barh(
                machine_index, bar_length, left=start_dt, color=color, label=detail
            )

            # min_dt = min(min_dt, start_dt)
            # max_dt = max(max_dt, end_dt)

        # Обновление пределов оси X
        self.ax.set_xlim([min_dt, max_dt])

        # Установка меток на оси X
        if range_selection == "День":
            self.ax.xaxis.set_major_locator(mdates.HourLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        elif range_selection == "Неделя":
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
        elif range_selection == "Месяц":
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))

        self.ax.set_xlabel("")
        self.ax.set_yticks(range(len(self.all_machines)))
        self.ax.set_yticklabels(self.all_machines)
        self.ax.xaxis.set_visible(True)
        self.ax.xaxis.grid(False)
        self.ax.yaxis.grid(False)
        self.ax.spines["top"].set_visible(False)
        self.ax.spines["right"].set_visible(False)
        self.ax.spines["left"].set_visible(False)
        self.ax.spines["bottom"].set_visible(False)
        self.ax.legend(loc="upper left", bbox_to_anchor=(1, 1), fontsize="small")
        plt.setp(
            self.ax.get_xticklabels(), rotation=90, ha="right", rotation_mode="anchor"
        )
        self.canvas.draw()

    def update_statusbar(self):
        """Обновляем статусбар с количеством записей и коротким путём"""
        num_tasks = len(self.task_table.get_children())
        num_items = len(self.nomenclature_table.get_children())
        if self.current_file_path == "":
            short_path = "Не задано"
        else:
            short_path = self.shorten_path(self.current_file_path)

        self.statusbar.config(
            text=f"Всего задач: {num_tasks}  | Всего записей в базе: {num_items} | База данных: {short_path}"
        )

    def update_detail_list(self):
        """Обновление списка деталей в комбобоксе"""
        self.detailList = self.nomenclature_table.get_column_values_by_index(1)
        self.detail_entry.set_completion_list(self.detailList)
        self.mark_data_as_modified()
        self.update_statusbar()

    def mark_data_as_modified(self, *args):
        self.is_data_modified = True

    def highlight_conflicts(self):
        if not self.task_table.get_children():
            return
        self.task_table.tag_configure("conflict", background="red", foreground="white")

        # Удаляем теги конфликтов перед повторной проверкой
        for item in self.task_table.get_children():
            self.task_table.item(item, tags="")

        machine_tasks = {}

        for item in self.task_table.get_children():
            (
                id,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            start_dt = datetime.strptime(start_datetime, "%d.%m.%Y %H:%M")
            end_dt = datetime.strptime(end_datetime, "%d.%m.%Y %H:%M")

            # Сохраняем задачи по машине
            if machine not in machine_tasks:
                machine_tasks[machine] = []
            machine_tasks[machine].append((item, start_dt, end_dt))

        # Проверяем конфликты только для каждой машины
        for machine, tasks in machine_tasks.items():
            for i in range(len(tasks)):
                item1, start1, end1 = tasks[i]
                for j in range(i + 1, len(tasks)):
                    item2, start2, end2 = tasks[j]
                    overlap_start = max(start1, start2)
                    overlap_end = min(end1, end2)
                    overlap_duration = (
                        overlap_end - overlap_start
                    ).total_seconds() / 60
                    if overlap_duration >= 0.5:  # Время в минутах
                        self.task_table.item(item1, tags="conflict")
                        self.task_table.item(item2, tags="conflict")

    def load_settings(self):
        """Загрузка настроек из JSON файла"""
        settings = {}
        settings_path = Path("settings.json").resolve()

        # Проверяем, существует ли файл и не пуст ли он
        if settings_path.exists() and settings_path.stat().st_size > 0:
            try:
                with settings_path.open("r", encoding="utf-8") as f:
                    settings = json.load(f)
            except json.JSONDecodeError as e:
                print(f"Ошибка декодирования JSON: {e}")
        else:
            print("Файл настроек не существует или пуст.")

        self.current_style = settings.get("style", "vista")
        self.change_style(self.current_style)
        self.csv_separator = settings.get("csv_separator", ";")
        self.encoding = settings.get("encoding", "windows-1251")
        self.all_machines = settings.get(
            "machines",
            [
                "HAAS VF-3",
                "DMU-50-1",
                "DMU-50-2",
                "DMU-70",
                "DMU-75",
                "DMG-M1-1",
                "DMG-M1-2",
                "DMG-M1-3",
                "DMC-835",
            ],
        )

        geometry = settings.get("window_geometry", None)
        if geometry:
            self.root.geometry(geometry)

        window_state = settings.get("window_state", "normal")
        self.root.state(window_state)

        self.current_file_path = settings.get("last_opened_file", "")
        self.save_settings()

    def save_settings(self):
        """Сохранение настроек в JSON файл"""
        geometry = self.root.geometry()
        window_state = self.root.state()

        if self.current_file_path and Path(self.current_file_path).exists():
            path = str(self.current_file_path)
        else:
            path = None

        settings = {
            "csv_separator": self.csv_separator,
            "encoding": self.encoding,
            "machines": self.all_machines,
            "style": self.current_style,
            "window_geometry": geometry,
            "window_state": window_state,
            "last_opened_file": path,
        }

        settings_path = Path("settings.json").resolve()
        with settings_path.open("w", encoding="utf-8") as f:
            json.dump(settings, f)

    def change_style(self, style_name):
        """Изменение стиля приложения"""
        style = ttk.Style()
        style.theme_use(style_name)
        self.current_style = style_name

    def setup_window(self):
        """Настройка главного окна"""
        self.root.title("Планирование")
        self.root.geometry("700x600")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_variables(self):
        """Настройка переменных"""
        self.is_data_modified = False
        # Создаем экземпляр валидатора
        self.validator = Validator(valid_values=self.get_valid_values())
        self.detailList = []
        self.settings_window = None
        self.about_window = None

    """Создание интерфейса"""

    def create_widgets(self):
        self.create_menu()
        self.create_statusbar()
        self.create_tabs()

        # Установка значений для комбобоксов
        self.machine_combo.set("DMU-50-1")
        self.machine_combo2.set("DMU-50-1")

    def create_menu(self):
        """Создание меню"""
        self.menu = tk.Menu(self.root)
        self.root.config(menu=self.menu)

        # Меню Файл
        file_menu = tk.Menu(self.menu, tearoff=0)
        file_menu.add_command(
            label="Открыть базу данных",
            command=self.open_database,
            accelerator="Ctrl+O",
        )
        file_menu.add_command(
            label="Сохранить базу данных",
            command=self.save_to_database,
            accelerator="Ctrl+S",
        )
        file_menu.add_command(label="Закрыть базу данных", command=self.close_database)
        file_menu.add_separator()
        file_menu.add_command(label="Импорт задач из Excel", command=self.import_tasks)
        file_menu.add_command(label="Экспорт задач в Excel", command=self.export_tasks)
        file_menu.add_command(
            label="Экспорт диаграммы в Excel", command=self.export_diagram
        )
        file_menu.add_separator()
        file_menu.add_command(label="Настройки", command=self.open_settings)
        file_menu.add_separator()
        file_menu.add_command(
            label="Выход", command=self.on_close, accelerator="Ctrl+Q"
        )
        self.menu.add_cascade(label="Файл", menu=file_menu)

        # Меню Правка
        edit_menu = tk.Menu(self.menu, tearoff=0)
        edit_menu.add_command(
            label="Вырезать", command=self.cmd_cut, accelerator="Ctrl+X"
        )
        edit_menu.add_command(
            label="Копировать", command=self.cmd_copy, accelerator="Ctrl+C"
        )
        edit_menu.add_command(
            label="Вставить", command=self.cmd_paste, accelerator="Ctrl+V"
        )
        edit_menu.add_command(
            label="Выделить все", command=self.cmd_select_all, accelerator="Ctrl+A"
        )
        self.menu.add_cascade(label="Правка", menu=edit_menu)

        # Меню Помощь
        help_menu = tk.Menu(self.menu, tearoff=0)
        help_menu.add_command(
            label="О программе", command=self.show_about, accelerator="F1"
        )
        help_menu.add_command(
            label="Руководство пользователя", command=self.open_user_guide
        )
        self.menu.add_cascade(label="Помощь", menu=help_menu)

        # Привязка горячих клавиш
        self.root.bind_all("<Control-o>", lambda event: self.open_database())
        self.root.bind_all("<Control-s>", lambda event: self.save_to_database())
        self.root.bind_all("<Control-q>", lambda event: self.on_close())
        self.root.bind_all("<F1>", lambda event: self.show_about())
        root.bind("<Control-KeyPress>", self.keypress)

    def keypress(self, e):
        """Обработчик комбинаций клавиш для вставки, копирования и вырезания"""
        if e.keycode == 86 and e.keysym != "v":
            self.cmd_paste()
        elif e.keycode == 67 and e.keysym != "c":
            self.cmd_copy()
        elif e.keycode == 88 and e.keysym != "x":
            self.cmd_cut()
        elif e.keycode == 65 and e.keysym != "a":
            self.cmd_select_all()

    def cmd_copy(self):
        """Обработчик команды копирования"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Copy>>")

    def cmd_cut(self):
        """Обработчик команды вырезания"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Cut>>")

    def cmd_paste(self):
        """Обработчик команды вставки"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<Paste>>")

    def cmd_select_all(self):
        """Обработчик команды выделения всего текста"""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Entry) or isinstance(widget, tk.Text):
            widget.event_generate("<<SelectAll>>")

    def create_tabs(self):
        """Создание вкладок"""
        self.tab_control = ttk.Notebook(self.root)
        self.tab_planning = ttk.Frame(self.tab_control)
        self.tab_diagram = ttk.Frame(self.tab_control)
        self.tab_nomenclature = ttk.Frame(self.tab_control)
        self.current_item = None
        self.tab_control.add(self.tab_planning, text="Планирование")
        self.tab_control.add(self.tab_diagram, text="Диаграмма")
        self.tab_control.add(self.tab_nomenclature, text="Номенклатура")
        self.tab_control.pack(expand=1, fill="both")

        self.create_planning_tab()
        self.create_diagram_tab()
        self.create_nomenclature_tab()

    def create_statusbar(self):
        """Создание статусбара"""
        self.statusbar = ttk.Label(self.root, text="", anchor=tk.W)
        self.statusbar.pack(side=tk.BOTTOM, fill=tk.X)

    def create_planning_tab(self):
        """Вкладка Планирование"""
        frame = ttk.Frame(self.tab_planning)
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Настройка ширины колонок
        frame.grid_columnconfigure(0, minsize=20)
        frame.grid_columnconfigure(1, minsize=50)

        # Поля ввода
        tk.Label(frame, text="Деталь").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        # self.detail_entry = tk.Entry(frame, width=100)
        self.detail_entry = AutoCompleteEntry(frame, width=100)
        self.detail_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Установ").grid(
            row=1, column=0, padx=5, pady=5, sticky="w"
        )
        self.setup_entry = tk.Entry(frame, width=13)
        self.setup_entry.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "setup")
        )
        self.setup_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Станок").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.machine_combo = ttk.Combobox(frame, width=10, values=self.all_machines)
        self.machine_combo.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Количество, шт").grid(
            row=3, column=0, padx=5, pady=5, sticky="w"
        )
        self.quantity_entry = tk.Entry(frame, width=13)
        self.quantity_entry.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "quantity")
        )
        self.quantity_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Время на 1 шт, мин").grid(
            row=4, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_entry = tk.Entry(frame, width=13)
        self.time_entry.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "time")
        )
        self.time_entry.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Время на наладку партии, мин").grid(
            row=5, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_setup = tk.Entry(frame, width=13)
        self.time_setup.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "time")
        )
        self.time_setup.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Дата запуска").grid(
            row=6, column=0, padx=5, pady=5, sticky="w"
        )
        self.start_date = DateEntry(frame, width=10, date_pattern="dd.MM.yyyy")
        self.start_date.grid(row=6, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Время запуска").grid(
            row=7, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_selector = TimeSelector(frame)
        self.time_selector.frame.grid(row=7, column=1, padx=5, pady=5, sticky="w")

        # Кнопки управления задачами
        control_frame = ttk.Frame(frame)
        control_frame.grid(row=8, columnspan=3, pady=10)
        self.add_task_button = tk.Button(
            control_frame, width=15, text="Запрос в базу", command=self.add_time
        )
        self.add_task_button.pack(side="left", padx=5)
        self.add_task_button = tk.Button(
            control_frame, width=15, text="Добавить", command=self.add_task
        )
        self.add_task_button.pack(side="left", padx=5)
        self.edit_task_button = tk.Button(
            control_frame, width=15, text="Редактировать", command=self.edit_task
        )
        self.edit_task_button.pack(side="left", padx=5)
        self.delete_task_button = tk.Button(
            control_frame, width=15, text="Удалить", command=self.delete_task
        )
        self.delete_task_button.pack(side="left", padx=5)
        self.clear_tasks_button = tk.Button(
            control_frame, width=15, text="Очистить", command=self.clear_tasks
        )
        self.clear_tasks_button.pack(side="left", padx=5)

        # Таблица
        self.task_table = EditableTreeview(
            frame,
            columns=[
                "ID",
                "Деталь",
                "Уст",
                "Станок",
                "Кол-во",
                "Время/шт",
                "Дата запуска",
                "Дата окончания",
            ],
            valid_values=self.get_valid_values(),
            update_app=self.update_app,
        )
        self.task_table.grid(row=9, columnspan=3, pady=10, sticky="nsew")
        self.task_table.column(
            "ID", anchor="center", width=1, minwidth=1, stretch=False
        )
        self.task_table.column(
            "Деталь", anchor="center", width=150, minwidth=100, stretch=True
        )
        self.task_table.column(
            "Уст", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.task_table.column(
            "Станок", anchor="center", width=50, minwidth=50, stretch=True
        )
        self.task_table.column(
            "Кол-во", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.task_table.column(
            "Время/шт", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.task_table.column(
            "Дата запуска", anchor="center", width=80, minwidth=80, stretch=True
        )
        self.task_table.column(
            "Дата окончания", anchor="center", width=80, minwidth=80, stretch=True
        )
        self.task_table.non_editable_columns = [
            "ID",
            "Кол-во",
            "Время/шт",
            "Дата запуска",
            "Дата окончания",
        ]

        # Скролбар
        scrollbar_y = ttk.Scrollbar(
            frame, orient="vertical", command=self.task_table.yview
        )
        scrollbar_y.grid(row=9, column=3, sticky="ns")
        self.task_table.configure(yscrollcommand=scrollbar_y.set)

        # Поле фильтрации
        self.filter_entry = tk.Entry(frame, width=30)
        self.filter_entry.grid(row=10, column=1, sticky="e")
        self.filter_entry.bind(
            "<KeyRelease>",
            lambda event: self.task_table.filter_rows(self.filter_entry.get()),
        )

        # Расширяем таблицу
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(9, weight=1)

    def create_diagram_tab(self):
        """Вкладка Диаграмма"""
        canvas_frame = ttk.Frame(self.tab_diagram)
        canvas_frame.pack(fill="both", expand=True)

        # Добавление выпадающего списка
        self.range_combo = ttk.Combobox(
            canvas_frame, values=["День", "Неделя", "Месяц"]
        )
        self.range_combo.pack(side="bottom", pady=10)
        self.range_combo.current(1)  # День по умолчанию
        self.range_combo.bind(
            "<<ComboboxSelected>>", self.update_gantt_chart
        )  # Обработка изменения выбора

        # Создание фигуры и осей для графика
        self.figure = Figure(figsize=(8, 6), dpi=100)
        self.ax = self.figure.add_subplot(111)

        # Создание канваса для графика
        self.canvas = FigureCanvasTkAgg(self.figure, master=canvas_frame)

        # Создание панели навигации
        self.toolbar = NavigationToolbar2Tk(self.canvas, canvas_frame)
        self.toolbar.update()
        self.toolbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.get_tk_widget().pack(
            side=tk.TOP, fill="both", expand=True
        )  # Центрирование канваса

        # Привязка события прокрутки
        self.canvas.mpl_connect("scroll_event", self.zoom)

        # Обновление графика
        self.update_gantt_chart()

    def create_nomenclature_tab(self):
        """Создание вкладки Номенклатура"""
        # Вкладка Номенклатура
        frame = ttk.Frame(self.tab_nomenclature)
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Настройка ширины колонок
        frame.grid_columnconfigure(0, minsize=20)
        frame.grid_columnconfigure(1, minsize=50)

        # Поля ввода
        tk.Label(frame, text="Тип детали").grid(
            row=0, column=0, padx=5, pady=5, sticky="w"
        )
        self.part_type_entry = tk.Entry(frame, width=30)
        self.part_type_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Номер чертежа").grid(
            row=1, column=0, padx=5, pady=5, sticky="w"
        )
        self.drawing_number_entry = tk.Entry(frame, width=30)
        self.drawing_number_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Установ").grid(
            row=3, column=0, padx=5, pady=5, sticky="w"
        )
        self.setup_entry2 = tk.Entry(frame, width=13)
        self.setup_entry2.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "setup")
        )
        self.setup_entry2.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Станок").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.machine_combo2 = ttk.Combobox(frame, width=10, values=self.all_machines)
        self.machine_combo2.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        tk.Label(frame, text="Время на 1 шт, мин").grid(
            row=5, column=0, padx=5, pady=5, sticky="w"
        )
        self.time_entry2 = tk.Entry(frame, width=13)
        self.time_entry2.bind(
            "<FocusOut>", lambda event: self.on_focus_out(event, "time")
        )
        self.time_entry2.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        # Кнопки управления номенклатурой
        control_frame = ttk.Frame(frame)
        control_frame.grid(row=6, columnspan=3, pady=10)
        self.add_nomenclature_button = tk.Button(
            control_frame, width=15, text="Добавить", command=self.add_nomenclature
        )
        self.add_nomenclature_button.pack(side="left", padx=5)
        self.edit_nomenclature_button = tk.Button(
            control_frame,
            width=15,
            text="Редактировать",
            command=self.edit_nomenclature,
        )
        self.edit_nomenclature_button.pack(side="left", padx=5)
        self.delete_nomenclature_button = tk.Button(
            control_frame, width=15, text="Удалить", command=self.delete_nomenclature
        )
        self.delete_nomenclature_button.pack(side="left", padx=5)

        # Таблица
        self.nomenclature_table = EditableTreeview(
            frame,
            columns=["ID", "Деталь", "Уст", "Станок", "Время/шт"],
            valid_values=self.get_valid_values(),
            update_app=self.update_app,
        )
        self.nomenclature_table.grid(row=7, columnspan=3, pady=10, sticky="nsew")
        self.nomenclature_table.column(
            "ID", anchor="center", width=5, minwidth=5, stretch=False
        )
        self.nomenclature_table.column(
            "Деталь", anchor="center", width=200, minwidth=200, stretch=True
        )
        self.nomenclature_table.column(
            "Уст", anchor="center", width=10, minwidth=10, stretch=True
        )
        self.nomenclature_table.column(
            "Станок", anchor="center", width=50, minwidth=50, stretch=True
        )
        self.nomenclature_table.column(
            "Время/шт", anchor="center", width=10, minwidth=10, stretch=True
        )

        # Скролбар
        scrollbar_y = ttk.Scrollbar(
            frame, orient="vertical", command=self.nomenclature_table.yview
        )
        scrollbar_y.grid(row=7, column=3, sticky="ns")
        self.nomenclature_table.configure(yscrollcommand=scrollbar_y.set)

        # Поле фильтрации
        self.filter_entry2 = tk.Entry(frame, width=30)
        self.filter_entry2.grid(row=8, column=2, sticky="e")
        self.filter_entry2.bind(
            "<KeyRelease>",
            lambda event: self.nomenclature_table.filter_rows(self.filter_entry2.get()),
        )

        # Расширяем таблицу
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(7, weight=1)

    def validate_field(self, field_type, value):
        """Валидация для полей при потере фокуса"""
        if field_type in ["setup", "quantity"]:
            col = "Уст" if field_type == "setup" else "Кол-во"
            return self.validator.validate_value(col, value)
        elif field_type == "time":
            return self.validator.validate_value("Время/шт", value)

        return True

    def on_focus_out(self, event, field_type):
        """Обработчик события потери фокуса для валидации"""
        value = event.widget.get()
        self.validate_field(field_type, value)

    def get_valid_values(self):
        return {
            "Уст": "positive_integer",  # Положительные целые
            "Станок": self.all_machines,  # Список
            "Кол-во": "positive_integer",
            "Время/шт": "positive_decimal",  # Положительные с точкой
            "Дата запуска": "datetime_format",  # Дата и время
            "Дата окончания": "datetime_format",
        }

    def validate_and_import_data(self, data):
        """Валидация данных перед загрузкой в таблицу"""
        valid_data = []
        errors = []

        for i, row in enumerate(data):
            try:
                # Валидация 1-й колонки (должна быть целым числом)
                int(row[0])

                # Валидация 5-й колонки (замена запятой и проверка на float)
                if len(row) > 4:
                    row[4] = row[4].replace(",", ".")
                    float(row[4])

                # Если валидация успешна, добавляем строку в валидные данные
                valid_data.append(row)
            except Exception as e:
                # Если ошибка — добавляем информацию в список ошибок
                errors.append((i + 1, str(e)))

        return valid_data, errors

    def load_database_silently(self, file_path=None, delimiter=","):
        """Загрузка базы данных без отображения ошибок"""
        if not file_path:
            return

        file_path = Path(file_path) if isinstance(file_path, str) else file_path
        data = []
        try:
            with file_path.open(newline="", encoding=self.encoding) as csvfile:
                reader = csv.reader(csvfile, delimiter=delimiter)
                next(reader)  # Пропускаем заголовок
                for row in reader:
                    if row:
                        data.append(row)
        except FileNotFoundError:
            print(f"Файл {file_path} не найден.")
            return
        except Exception as e:
            print(f"Ошибка при загрузке файла: {e}")
            return

        # Проверка и валидация данных перед загрузкой в таблицу
        valid_data, errors = self.validate_and_import_data(data)

        # Вывод ошибок в консоль
        if errors:
            for error in errors:
                print(f"Ошибка в строке {error[0]}: {error[1]}")

        if not valid_data:
            print("Все строки содержат ошибки.")
            return

        # Загрузка валидных данных в TreeView
        self.nomenclature_table.clear_data()
        self.nomenclature_table.insert_data(valid_data)
        self.update_detail_list()
        self.is_data_modified = False
        print(f"Успешно загружено {len(valid_data)} записей из {file_path}.")

    def open_database(self):
        """Открытие базы данных из CSV файла"""
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not file_path:
            return  # Если пользователь отменил выбор файла
        file_path = Path(file_path)

        if self.current_file_path == file_path:
            messagebox.showwarning("Предупреждение", "Файл уже открыт.")
            return

        if not self.csv_separator:
            delimiter = simpledialog.askstring(
                "Выбор разделителя", "Введите разделитель (например, ',' или ';'):"
            )
        else:
            delimiter = self.csv_separator

        if not delimiter:
            return  # Если пользователь не ввел разделитель

        data = []
        try:
            with file_path.open(newline="", encoding=self.encoding) as csvfile:
                reader = csv.reader(csvfile, delimiter=delimiter)
                header = next(reader)  # Пропускаем первую строку (заголовок)
                for row in reader:
                    if row:  # Проверка на пустую строку
                        data.append(row)
        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Указанный файл не найден.")
            return
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {e}")
            return

        if not data:
            messagebox.showwarning(
                "Предупреждение", "Файл пустой или неправильно отформатирован."
            )
            return

        # Валидация данных
        valid_data, errors = self.validate_and_import_data(data)

        # Обработка ошибок валидации
        if errors:
            error_message = "\n".join(
                [f"Ошибка в строке {i}: {error}" for i, error in errors]
            )
            messagebox.showerror(
                "Ошибки при импорте", f"Обнаружены следующие ошибки:\n{error_message}"
            )
            return  # Прекращаем выполнение, если есть ошибки

        # Если валидация прошла, очищаем таблицу и добавляем валидные данные
        self.nomenclature_table.clear_data()
        self.tab_control.select(self.tab_nomenclature)
        self.nomenclature_table.insert_data(valid_data)
        self.current_file_path = file_path
        self.update_detail_list()
        messagebox.showinfo(
            "Успех", f"Успешно импортировано {len(valid_data)} записей."
        )

    def export_to_csv(self, file_path):
        """Экспорт данных из Treeview в CSV файл"""
        file_path = Path(file_path)

        if not self.csv_separator:
            delimiter = simpledialog.askstring(
                "Выбор разделителя", "Введите разделитель (например, ',' или ';'):"
            )
        else:
            delimiter = self.csv_separator

        if not delimiter:
            return  # Если пользователь не ввел разделитель
        try:
            with file_path.open("w", newline="", encoding=self.encoding) as csvfile:
                writer = csv.writer(csvfile, delimiter=delimiter)

                # Запись заголовков (если необходимо)
                writer.writerow(self.nomenclature_table["columns"])

                # Запись данных из Treeview
                for row in self.nomenclature_table.get_children():
                    writer.writerow(self.nomenclature_table.item(row)["values"])

            self.current_file_path = file_path
            return True  # Успешный экспорт
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать данные: {e}")
            return False  # Неуспешный экспорт

    def save_to_database(self):
        """Сохранение данных из Treeview в CSV файл"""
        if self.current_file_path and self.current_file_path.exists():
            success = self.export_to_csv(self.current_file_path)
        else:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv", filetypes=[("CSV files", "*.csv")]
            )
            if file_path:
                success = self.export_to_csv(file_path)
            else:
                return  # Если пользователь отменил выбор файла, выходим из функции

        if success:
            messagebox.showinfo("Успех", "Файл успешно сохранен.")
            self.update_detail_list()

    def close_database(self):
        """Закрытие базы данных"""
        for item in self.nomenclature_table.get_children():
            self.nomenclature_table.delete(item)
        self.current_file_path = ""
        self.update_detail_list()

    def import_tasks(self):
        file_path = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return

        file_path = Path(file_path)
        try:
            # Загружаем данные из Excel без преобразования дат
            df = pd.read_excel(file_path, dtype=str)

            if df.empty:
                messagebox.showwarning("Ошибка", "Файл пустой.")
                return

            self.task_table.clear_data()

            for _, row in df.iterrows():
                # Получаем значения из строк как есть
                start_date = row.get("Дата запуска", "")
                end_date = row.get("Дата окончания", "")

                task_data = [
                    row.get("Деталь", ""),
                    row.get("Установ", ""),
                    row.get("Станок", ""),
                    row.get("Кол-во", ""),
                    row.get("Время/шт", ""),
                    start_date,
                    end_date,
                ]
                self.task_table.add_row(task_data)

            self.update_app()
            self.is_data_modified = False
            messagebox.showinfo("Успех", "Задачи успешно загружены.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при загрузке задач: {str(e)}")

    def save_to_excel(
        self, file_path, data, sheet_name, include_image=None, column_names=None
    ):
        """Сохранение данных в Excel файл"""
        file_path = Path(file_path)
        try:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Если названия колонок не переданы, используем индексы по умолчанию
                df = (
                    pd.DataFrame(data, columns=column_names)
                    if column_names
                    else pd.DataFrame(data)
                )
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Добавляем отступ между заголовком и данными
                # worksheet.insert_rows(1)
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=1, column=col_num).alignment = Alignment(
                        horizontal="center"
                    )

                # Установка ширины колонок
                for column in worksheet.columns:
                    max_length = 0
                    column = [
                        cell for cell in column
                    ]  # Преобразование в список для обработки
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (
                        max_length + 2
                    )  # Увеличиваем ширину на 2 для отступа
                    worksheet.column_dimensions[column[0].column_letter].width = (
                        adjusted_width
                    )

                # Центрируем содержимое всех ячеек
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=worksheet.max_row,
                    min_col=1,
                    max_col=worksheet.max_column,
                ):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

                # Если передано изображение, добавляем его в Excel
                if include_image:
                    img = Image(include_image)
                    worksheet.add_image(
                        img, "H1"
                    )  # Замените "H1" на нужную ячейку для изображения

            messagebox.showinfo("Успех", "Данные успешно сохранены.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при сохранении в Excel: {str(e)}")

    def export_tasks(self):
        """Экспорт задач в Excel файл"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return

        file_path = Path(file_path)
        data = []
        for item in self.task_table.get_children():
            (
                _,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            data.append(
                (
                    detail,
                    setup,
                    machine,
                    quantity,
                    time_per_unit,
                    start_datetime,
                    end_datetime,
                )
            )

        # Передаем названия колонок
        column_names = [
            "Деталь",
            "Установ",
            "Станок",
            "Кол-во",
            "Время/шт",
            "Дата запуска",
            "Дата окончания",
        ]
        self.save_to_excel(
            file_path, data, sheet_name="Tasks", column_names=column_names
        )

    def export_diagram(self):
        """Экспорт диаграммы в Excel файл"""
        excel_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Сохранить как",
        )
        if not excel_file:
            messagebox.showwarning("Внимание", "Файл не выбран. Экспорт отменен.")
            return

        excel_file = Path(excel_file)
        image_path = "gantt_chart.png"
        self.figure.savefig(image_path, bbox_inches="tight")
        plt.close()

        data = []
        for item in self.task_table.get_children():
            (
                _,
                detail,
                setup,
                machine,
                quantity,
                time_per_unit,
                start_datetime,
                end_datetime,
            ) = self.task_table.item(item, "values")
            data.append(
                [
                    detail,
                    setup,
                    machine,
                    quantity,
                    time_per_unit,
                    start_datetime,
                    end_datetime,
                ]
            )

        # Передаем названия колонок
        column_names = [
            "Деталь",
            "Установ",
            "Станок",
            "Кол-во",
            "Время/шт",
            "Дата запуска",
            "Дата окончания",
        ]
        self.save_to_excel(
            excel_file,
            data,
            sheet_name="Diagram",
            include_image=image_path,
            column_names=column_names,
        )

    def open_settings(self):
        """Открытие окна настроек"""
        if self.settings_window is None or not self.settings_window.winfo_exists():
            self.settings_window = SettingsWindow(self.root, self)
        else:
            self.settings_window.lift()

    def on_close(self):
        """Обработчик закрытия окна"""
        self.save_settings()
        if self.is_data_modified:
            if not messagebox.askokcancel(
                "Выход",
                "Есть несохранённые изменения в базе данных. Вы действительно хотите выйти?",
            ):
                return
        self.root.quit()
        self.root.destroy()

    def show_about(self):
        """О программе"""
        if self.about_window is None or not self.about_window.winfo_exists():
            self.about_window = tk.Toplevel()
            self.about_window.title("О программе")
            self.about_window.geometry("300x200")
            self.about_window.resizable(False, False)

            # Запретить сворачивание окна
            self.about_window.wm_attributes("-toolwindow", True)

            app_info = [
                ("Название: ", "bold"),
                ("Планирование\n", "normal"),
                ("Версия: ", "bold"),
                (f"{__version__}\n", "normal"),
                ("Автор: ", "bold"),
                ("MaestroFusion360\n", "normal"),
                ("Описание: ", "bold"),
                (
                    "Это приложение позволяет управлять производственными задачами, "
                    "планировать их выполнение и визуализировать в виде диаграммы Ганта.\n",
                    "normal",
                ),
            ]

            text_widget = tk.Text(self.about_window, wrap="word", height=10, width=40)
            text_widget.pack(padx=10, pady=10, expand=True, fill="both")

            for text, style in app_info:
                if style == "bold":
                    text_widget.insert(tk.END, text, ("bold",))
                else:
                    text_widget.insert(tk.END, text)

            text_widget.tag_config("bold", font=("Arial", 10, "bold"))
            text_widget.config(state=tk.DISABLED)

            close_button = ttk.Button(
                self.about_window, text="Закрыть", command=self.about_window.destroy
            )
            close_button.pack(pady=10)

        else:
            self.about_window.lift()

    def open_user_guide(self):
        """Открытие руководства пользователя"""
        # Получаем путь к текущему скрипту и строим путь к файлу README.md
        current_directory = Path(__file__).parent
        file_path = f"{current_directory}/README_RU.md"

        if Path(file_path).exists():
            # subprocess.Popen(["notepad.exe", str(file_path)])  # Открыть в Notepad
            self.open_markdown_in_browser(file_path)  # Открыть в браузере
        else:
            messagebox.showerror("Ошибка", "Файл руководства не найден.")

    def add_time(self):
        """Запрос времени на деталь"""
        detail = self.detail_entry.get().strip()
        setup = self.setup_entry.get().strip()
        machine = self.machine_combo.get().strip()
        search_row = (detail, setup, machine)

        # Проверка на пустые поля
        if not all([detail, setup, machine]):
            messagebox.showwarning(
                "Внимание", "Пожалуйста, заполните эти поля: Деталь, Установ, Станок"
            )
            return

        for child in self.nomenclature_table.get_children():
            row_values = self.nomenclature_table.item(child, "values")

            if (
                row_values[1] == search_row[0]
                and row_values[2] == search_row[1]
                and row_values[3] == search_row[2]
            ):
                messagebox.showinfo(
                    "Результат поиска", f"Найдена номенклатура: {row_values}"
                )
                self.time_entry.delete(0, tk.END)
                self.time_entry.insert(0, row_values[4])
                return

        messagebox.showwarning("Результат поиска", "Номенклатура не найдена.")

    def add_task(self):
        """Добавление задачи"""
        detail = self.detail_entry.get().strip()
        setup = self.setup_entry.get().strip()
        machine = self.machine_combo.get().strip()
        quantity = self.quantity_entry.get().strip()
        time_per_unit = self.time_entry.get().replace(",", ".")
        time_setup = self.time_setup.get().replace(",", ".")
        start_date_str = self.start_date.get()
        start_time_str = self.time_selector.get_time()

        # Проверка на пустые поля
        if not all([detail, setup, machine, quantity, time_per_unit]):
            messagebox.showwarning("Внимание", "Пожалуйста, заполните все поля.")
            return

        try:
            duration = (
                int(quantity) * float(time_per_unit) / 60 + float(time_setup) / 60
            )
            start_datetime = datetime.strptime(
                f"{start_date_str} {start_time_str}", "%d.%m.%Y %H:%M"
            )
            end_datetime = start_datetime + timedelta(hours=duration)

            # Добавление задачи в таблицу
            self.task_table.add_row(
                [
                    detail,
                    setup,
                    machine,
                    quantity,
                    time_per_unit,
                    start_datetime.strftime("%d.%m.%Y %H:%M"),
                    end_datetime.strftime("%d.%m.%Y %H:%M"),
                ]
            )

            self.update_app()
        except ValueError as ve:
            messagebox.showerror("Ошибка", f"Неверный ввод: {str(ve)}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

    def edit_task(self):
        """Редактирование задачи"""
        self.task_table.on_edit_row()
        self.update_app()

    def delete_task(self):
        """Удаление задачи"""
        self.task_table.delete_row()
        self.update_app()

    def clear_tasks(self):
        """Очистка задач"""
        for item in self.task_table.get_children():
            self.task_table.delete(item)
        self.update_app()

    def add_nomenclature(self):
        """Добавление номенклатуры"""
        type_detail = self.part_type_entry.get().strip()
        drawing_number = self.drawing_number_entry.get().strip()
        setup = self.setup_entry2.get().strip()
        machine = self.machine_combo2.get().strip()
        time_per_unit = self.time_entry2.get().strip().replace(",", ".")

        if type_detail and drawing_number and setup and machine and time_per_unit:
            detail = f"{type_detail} {drawing_number}"

            # Проверяем наличие записи с такими же detail, setup и machine
            for row_id in self.nomenclature_table.get_children():
                values = self.nomenclature_table.item(row_id, "values")
                existing_detail = values[1].strip()
                existing_setup = values[2].strip()
                existing_machine = values[3].strip()

                if (
                    existing_detail == detail
                    and existing_setup == setup
                    and existing_machine == machine
                ):
                    messagebox.showinfo(
                        "Информация",
                        f"Запись с такими данными уже существует под номером {self.nomenclature_table.index(row_id) + 1}",
                    )

                    # Выделяем существующую запись
                    self.nomenclature_table.selection_set(row_id)
                    self.nomenclature_table.see(row_id)
                    return

            # Если такой записи нет, добавляем новую строку
            self.nomenclature_table.add_row([detail, setup, machine, time_per_unit])
            self.update_detail_list()
        else:
            messagebox.showwarning("Внимание", "Пожалуйста, заполните все поля.")

    def edit_nomenclature(self):
        """Редактирование номенклатуры"""
        self.nomenclature_table.on_edit_row()
        self.update_detail_list()

    def delete_nomenclature(self):
        """Удаление номенклатуры"""
        self.nomenclature_table.delete_row()
        self.update_detail_list()

    def zoom(self, event):
        """Обработка события прокрутки мыши для изменения масштаба графика"""
        xmin, xmax = self.ax.get_xlim()
        xmin_dt = mdates.num2date(xmin)
        xmax_dt = mdates.num2date(xmax)
        zoom_factor = 0.9 if event.step > 0 else 1.1
        range_x = (xmax - xmin) * zoom_factor
        midpoint_x = (xmax + xmin) / 2
        self.ax.set_xlim(midpoint_x - range_x / 2, midpoint_x + range_x / 2)
        new_range_days = (xmax_dt - xmin_dt).days
        if new_range_days <= 1:
            self.ax.xaxis.set_major_locator(mdates.HourLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        elif new_range_days <= 7:
            self.ax.xaxis.set_major_locator(mdates.HourLocator(interval=6))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m %H:%M"))
        elif new_range_days <= 31:
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
        else:
            self.ax.xaxis.set_major_locator(mdates.DayLocator(interval=5))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m.%Y"))
        self.canvas.draw_idle()

    def generate_colors(self, num_colors):
        """Генерация цветов в формате HSV"""
        return [
            mcolors.hsv_to_rgb((i / num_colors, 0.8, 0.8)) for i in range(num_colors)
        ]

    def get_monday_at_midnight(self, reference_date):
        """Возвращает дату понедельника этой недели с установленным временем на 0:00."""
        days_since_monday = reference_date.weekday()  # Понедельник - 0, Воскресенье - 6
        monday = reference_date - timedelta(days=days_since_monday)
        return monday.replace(hour=0, minute=0, second=0, microsecond=0)

    def shorten_path(self, path, max_length=50):
        """Сокращение пути до заданной длины, заменяя средние части на '...'"""
        if path is None:
            return ""

        path = Path(path)

        # Если длина пути короче или равна максимальной длине, возвращаем как есть
        if len(str(path)) <= max_length:
            return str(path)

        # Разбиваем путь на части
        parts = list(path.parts)

        # Если путь состоит из более чем двух частей, сокращаем
        if len(parts) > 2:
            # Обрезаем путь, заменяя средние части на '...'
            shortened = Path(parts[0], "...", parts[-1])
        else:
            # Если меньше или равно двум частям, возвращаем оригинальный путь
            shortened = path

        return str(shortened)

    def open_markdown_in_browser(self, md_path):
        """Открытие Markdown файла в браузере"""
        md_file = Path(md_path)
        html_file = md_file.with_suffix(".html")

        # Читаем содержимое Markdown
        md_text = md_file.read_text(encoding="utf-8")

        # Конвертируем в HTML
        html = markdown.markdown(md_text, extensions=["fenced_code", "tables"])

        # Добавляем базовую HTML-обёртку
        full_html = f"""
        <html>
        <head>
            <meta charset="utf-8">
            <title>{md_file.stem}</title>
            <style>
                body {{ font-family: sans-serif; margin: 40px; max-width: 800px; }}
                code {{ background-color: #eee; padding: 2px 4px; }}
                pre code {{ background-color: #f8f8f8; display: block; padding: 10px; }}
            </style>
        </head>
        <body>{html}</body>
        </html>
        """

        # Сохраняем как HTML
        html_file.write_text(full_html, encoding="utf-8")

        # Открываем в браузере
        webbrowser.open(html_file.as_uri())


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
